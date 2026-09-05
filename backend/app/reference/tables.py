"""Reference data, loaded the same way whatever it came from.

Resolution — is this counterparty a real one, does this project code exist — is
lookup against tables somebody else maintains. The tables arrive as spreadsheet
sheets today and could arrive as CSV or a database tomorrow, so this module
knows about *tables*, not about counterparties.

Two decisions worth stating, both learned from the data rather than chosen:

**Headers are normalised on the way in.** The supplied workbook has a sheet
named `'DIU '` and columns `'Value date '`, `'Post date '`, `'Account '` — all
with trailing spaces — and cells carrying leading tabs. Every consumer would
otherwise have to remember, and one of them eventually would not.

**Lookup is exact first, then casefolded, and never fuzzy.** A near match is a
*candidate* for a human, not an answer. The whole point of the three-state
design is that "no match" stays "no match": 52 of the 100 rows in this dataset
genuinely have no counterparty, and quietly resolving them to the nearest
master-list name is the single worst thing this pipeline could do.
"""

from __future__ import annotations

import difflib
import json
from dataclasses import dataclass, field
from pathlib import Path

ROOT = Path(__file__).resolve().parents[3]


def normalise(value: object) -> str:
    """One string form for a cell, so lookups are not defeated by whitespace.

    Leading tabs are real in this data: two `Resolved Position` cells begin
    with one, and they are why four positions appear not to resolve.
    """
    if value is None:
        return ""
    text = str(value)
    return " ".join(text.split())


@dataclass
class Table:
    """One reference list: named columns, rows as dicts, indexed for lookup."""

    name: str
    columns: list[str]
    rows: list[dict] = field(default_factory=list)
    _index: dict[str, dict[str, dict]] = field(default_factory=dict, repr=False)

    def _for(self, column: str) -> dict[str, dict]:
        if column not in self._index:
            if column not in self.columns:
                raise KeyError(
                    f"table {self.name!r} has no column {column!r} "
                    f"(has: {', '.join(self.columns)})"
                )
            built: dict[str, dict] = {}
            for row in self.rows:
                key = row.get(column, "")
                if key:
                    built.setdefault(key.casefold(), row)
            self._index[column] = built
        return self._index[column]

    def contains(self, column: str, value: str) -> bool:
        return normalise(value).casefold() in self._for(column)

    def find(self, column: str, value: str) -> dict | None:
        """The row this value names, or None. Exact, then case-insensitive."""
        return self._for(column).get(normalise(value).casefold())

    def values(self, column: str) -> list[str]:
        seen = {row[column] for row in self.rows if row.get(column)}
        return sorted(seen)

    def candidates(self, column: str, value: str, limit: int = 5) -> list[str]:
        """Near matches, ranked — for a human to choose between, never to apply.

        Offered so an unresolved row arrives as a decision rather than an
        investigation. Nothing in the pipeline may promote one of these to a
        match on its own.
        """
        return difflib.get_close_matches(
            normalise(value), self.values(column), n=limit, cutoff=0.6
        )

    def to_json(self) -> dict:
        return {"name": self.name, "columns": self.columns, "rows": self.rows}


def _clean_sheet(raw: list[tuple], header_row: int, keep: list[str] | None) -> Table | None:
    """Turn a sheet's cells into a table, dropping padding and blanks."""
    if len(raw) <= header_row:
        return None

    headers = [normalise(c) for c in raw[header_row]]
    # Columns with no header are padding: the Deal & Position master declares 18
    # and populates 11; the Staging Sheet declares 25 and populates 24.
    live = [(i, h) for i, h in enumerate(headers) if h]
    if keep:
        wanted = {k.casefold() for k in keep}
        live = [(i, h) for i, h in live if h.casefold() in wanted]
    if not live:
        return None

    rows = []
    for cells in raw[header_row + 1 :]:
        row = {h: normalise(cells[i]) if i < len(cells) else "" for i, h in live}
        if any(row.values()):
            rows.append(row)

    return Table(name="", columns=[h for _, h in live], rows=rows)


def from_workbook(path: Path, spec: dict) -> dict[str, Table]:
    """Load the sheets a profile asked for, and only the columns it named.

    Column selection is not tidiness: the deal and position master is 6,635
    rows, and everything loaded here is later handed to a sandbox. Carrying
    columns nobody reads makes every run slower for no gain.
    """
    import openpyxl

    book = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        by_normalised = {normalise(name): name for name in book.sheetnames}
        tables: dict[str, Table] = {}

        for alias, want in spec.items():
            sheet_name = want["sheet"]
            actual = by_normalised.get(normalise(sheet_name))
            if actual is None:
                raise KeyError(
                    f"no sheet {sheet_name!r} in {path.name} "
                    f"(has: {', '.join(book.sheetnames)})"
                )

            raw = list(book[actual].iter_rows(values_only=True))
            table = _clean_sheet(raw, want.get("header_row", 0), want.get("columns"))
            if table is None:
                raise ValueError(f"sheet {sheet_name!r} yielded no usable columns")
            table.name = alias
            tables[alias] = table

        return tables
    finally:
        book.close()


def resolve_source(location: str) -> Path:
    """Find a declared input, preferring the committed copy.

    Mirrors how `cli.py` resolves the statements: the organisers committed the
    dataset under its own folder, and some working copies still hold the older
    flat unpack. Both must work, and a directory that exists but is empty must
    not shadow one that has the file.
    """
    candidates = [ROOT / location]
    if location.startswith("samples/01-bank-statements-to-journal-entries/"):
        tail = location.split("/", 2)[2]
        candidates.append(ROOT / "samples" / tail)

    for candidate in candidates:
        if candidate.is_file():
            return candidate
        if candidate.is_dir():
            found = sorted(candidate.glob("*.xlsx"))
            if found:
                return found[0]
    raise FileNotFoundError(f"no reference source at {' or '.join(map(str, candidates))}")


def load_tables(inputs: dict) -> dict[str, Table]:
    """Every table a profile declares. Returns {} when it declares none."""
    spec = inputs.get("tables") or {}
    if not spec:
        return {}
    return from_workbook(resolve_source(inputs["workbook"]["location"]), spec)


def dump(tables: dict[str, Table], path: Path) -> Path:
    """Write the tables where a sandbox can read them without openpyxl.

    The sandbox has pdfplumber and nothing else, deliberately — it runs
    model-written code. Serialising here keeps it that way.
    """
    payload = {name: table.to_json() for name, table in tables.items()}
    path.write_text(json.dumps(payload), encoding="utf-8")
    return path
