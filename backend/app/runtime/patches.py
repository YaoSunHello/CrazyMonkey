"""Verified, source-bound workbook corrections with an explicit audit trail.

Only proposals minted by ``propose_patch`` in this process can be applied. A
serialized/model-authored proposal is a review artifact, never write authority.
Original workbooks remain untouched and publication never replaces an existing
path, including when another process creates that path after preflight.
"""
from __future__ import annotations

import hashlib
import json
import os
import tempfile
import warnings
from collections import defaultdict
from copy import copy
from dataclasses import dataclass
from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from pydantic import BaseModel, ConfigDict, Field, PrivateAttr, field_validator

from app.atlas.ids import sha256_bytes
from app.atlas.ingestion import MAX_FILE_BYTES
from .challenger import challenge
from .contracts import Challenge, VerificationPlan
from .executor import execute
from .investigation_evidence import EvidenceStore


TRAIL_NAME = "Audit Trail"
TRAIL_HEADERS = ("Timestamp", "Sheet", "Cell", "Old Value", "New Value", "Finding", "Reason", "Evidence")


@dataclass(frozen=True)
class _VerificationProof:
    proposal_digest: str
    plan_json: str
    calculation_json: str
    document_hash: str


class PatchProposal(BaseModel):
    model_config = ConfigDict(extra="forbid", frozen=True)
    source_file: str = Field(min_length=1, max_length=255)
    document_id: str = Field(min_length=1, max_length=100)
    sheet: str = Field(min_length=1, max_length=31)
    cell: str = Field(pattern=r"^[A-Z]{1,3}[1-9][0-9]{0,6}$")
    old_value: str = Field(min_length=1, max_length=32767)
    new_value: str = Field(min_length=1, max_length=80)
    reason: str = Field(min_length=1, max_length=3000)
    evidence_ids: list[str] = Field(min_length=1, max_length=217)
    finding_id: str = Field(min_length=1, max_length=100)
    _proof: _VerificationProof | None = PrivateAttr(default=None)

    @field_validator("new_value")
    @classmethod
    def valid_amount(cls, value: str) -> str:
        _excel_number(value)
        return value


def _digest(proposal: PatchProposal) -> str:
    data = json.dumps(proposal.model_dump(mode="json"), sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(data.encode("utf-8")).hexdigest()


def _excel_number(value: str) -> int | float:
    try:
        amount = Decimal(value)
        if not amount.is_finite() or amount != amount.quantize(Decimal("0.01")):
            raise ValueError("Patch amount must be finite and exact to a penny")
        if len(amount.normalize().as_tuple().digits) > 15:
            raise ValueError("Patch amount exceeds Excel's 15 significant digit limit")
        numeric = int(amount) if amount == amount.to_integral_value() else float(amount)
        if Decimal(str(numeric)) != amount:
            raise ValueError("Patch amount cannot be represented exactly by the workbook writer")
        return numeric
    except (InvalidOperation, OverflowError) as exc:
        raise ValueError("Patch amount is outside supported numeric bounds") from exc


def propose_patch(plan: VerificationPlan, calculation: dict, store: EvidenceStore,
                  red_team: Challenge) -> PatchProposal | None:
    """Return a correction only for a verified discrepancy at one original cell.

    Unsupported/ambiguous targets and unaccepted findings return None. Invalid
    source references, modified originals and inconsistent calculations raise.
    The caller supplies the final verified red-team Challenge, never model prose.
    """
    plan = VerificationPlan.model_validate(plan.model_dump())
    red_team = Challenge.model_validate(red_team.model_dump())
    if (calculation.get("status") != "DISCREPANCY" or red_team.status != "PASS"
            or not red_team.checks or not all(red_team.checks.values())):
        return None
    store.verify_originals()
    recalculated = execute(plan, store, Decimal(calculation["tolerance"]))
    if recalculated != calculation:
        raise ValueError("Patch calculation does not equal independent source execution")
    if (plan.check_type != "model_proposed"
            and challenge(plan, recalculated, store, Decimal(calculation["tolerance"])).status != "PASS"):
        return None
    ref = store.get(plan.inputs[plan.reported_input].evidence_id)
    if (ref.kind != "WORKBOOK_CELL" or ref.formula or ref.cache_status != "NOT_APPLICABLE"
            or not ref.sheet or not ref.cell or ref.sheet.casefold() == TRAIL_NAME.casefold()):
        return None
    document = store.docs[ref.document_id].document
    if Path(document.filename).suffix.lower() != ".xlsx":
        return None
    matches = [item for item in store.refs.values()
               if item.document_id == ref.document_id and item.sheet == ref.sheet and item.cell == ref.cell]
    if len(matches) != 1 or ref.original_value is None:
        return None
    evidence_ids = list(dict.fromkeys(
        [item.evidence_id for item in plan.inputs.values()] + plan.context_evidence_ids))
    for evidence_id in evidence_ids:
        store.get(evidence_id)
    proposal = PatchProposal(
        source_file=document.filename, document_id=ref.document_id,
        sheet=ref.sheet, cell=ref.cell, old_value=ref.original_value,
        new_value=recalculated["expected"], reason=plan.rationale,
        evidence_ids=evidence_ids, finding_id=plan.check_id,
    )
    proposal._proof = _VerificationProof(
        proposal_digest=_digest(proposal), plan_json=plan.model_dump_json(),
        calculation_json=json.dumps(recalculated, sort_keys=True),
        document_hash=document.document_hash,
    )
    return proposal


def _file_hash(path: Path) -> str:
    with path.open("rb") as source:
        data = source.read(MAX_FILE_BYTES + 1)
    if len(data) > MAX_FILE_BYTES:
        raise ValueError("Workbook exceeds the supported source-size bound")
    return sha256_bytes(data)


def _validate_proposal(proposal: PatchProposal, store: EvidenceStore) -> Path:
    proof = proposal._proof
    if proof is None or _digest(proposal) != proof.proposal_digest:
        raise ValueError("Patch was not created by verified execution or has been modified")
    document = store.docs.get(proposal.document_id)
    if document is None or document.document.document_hash != proof.document_hash:
        raise ValueError("Patch document does not match its verified source")
    if proposal.source_file != document.document.filename:
        raise ValueError("Patch filename does not match its verified source")
    plan = VerificationPlan.model_validate_json(proof.plan_json)
    expected = json.loads(proof.calculation_json)
    if execute(plan, store, Decimal(expected["tolerance"])) != expected:
        raise ValueError("Patch source calculation changed after verification")
    ref = store.get(plan.inputs[plan.reported_input].evidence_id)
    if (ref.document_id, ref.sheet, ref.cell, ref.original_value) != (
            proposal.document_id, proposal.sheet, proposal.cell, proposal.old_value):
        raise ValueError("Patch destination does not match the reported evidence")
    if proposal.new_value != expected["expected"]:
        raise ValueError("Patch amount does not match the verified expected amount")
    for evidence_id in proposal.evidence_ids:
        store.get(evidence_id)
    return Path(document.document.original_storage_key).resolve()


def _cell_state(cell) -> tuple:
    hyperlink = cell.hyperlink
    comment = cell.comment
    return (cell.value, cell.data_type, copy(cell._style),
            (hyperlink.target, hyperlink.location, hyperlink.display, hyperlink.tooltip) if hyperlink else None,
            (comment.text, comment.author) if comment else None)


def _assert_preserved(original, corrected, proposals: list[PatchProposal]) -> None:
    expected_names = list(original.sheetnames)
    if TRAIL_NAME not in expected_names:
        expected_names.append(TRAIL_NAME)
    if corrected.sheetnames != expected_names:
        raise ValueError("Workbook sheet structure changed outside the audit trail")
    targets = {(item.sheet, item.cell): Decimal(item.new_value) for item in proposals}
    for old_sheet in original.worksheets:
        new_sheet = corrected[old_sheet.title]
        if (old_sheet.sheet_state != new_sheet.sheet_state
                or old_sheet.freeze_panes != new_sheet.freeze_panes
                or str(old_sheet.merged_cells) != str(new_sheet.merged_cells)):
            raise ValueError("Workbook sheet presentation changed unexpectedly")
        for old_cell in old_sheet._cells.values():
            new_cell = new_sheet[old_cell.coordinate]
            before, after = _cell_state(old_cell), _cell_state(new_cell)
            target = targets.get((old_sheet.title, old_cell.coordinate))
            if target is None:
                if before != after:
                    raise ValueError("Unrelated cell content or style changed during workbook save")
            elif (new_cell.data_type != "n" or Decimal(str(new_cell.value)) != target
                    or before[2:] != after[2:]):
                raise ValueError("Saved patch amount or original cell style did not survive exactly")


def _audit_trail(workbook, proposals: list[PatchProposal], timestamp: str) -> None:
    matching = [name for name in workbook.sheetnames if name.casefold() == TRAIL_NAME.casefold()]
    if matching and matching != [TRAIL_NAME]:
        raise ValueError("Existing audit-trail worksheet name is ambiguous")
    if matching:
        sheet = workbook[TRAIL_NAME]
        if sheet.max_column != len(TRAIL_HEADERS) or tuple(cell.value for cell in sheet[1]) != TRAIL_HEADERS:
            raise ValueError("Existing Audit Trail does not have the required column headers")
    else:
        sheet = workbook.create_sheet(TRAIL_NAME)
        sheet.append(TRAIL_HEADERS)
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        for column, width in zip("ABCDEFGH", (27, 25, 12, 20, 20, 28, 65, 90)):
            sheet.column_dimensions[column].width = width
        sheet.freeze_panes = "A2"
    for proposal in proposals:
        row = (timestamp, proposal.sheet, proposal.cell, proposal.old_value, proposal.new_value,
               proposal.finding_id, proposal.reason, ", ".join(proposal.evidence_ids))
        sheet.append(row)
        # All fields are recorded as exact text. Never turn model/source text
        # beginning with '=' into an executable workbook formula.
        for cell in sheet[sheet.max_row]:
            cell.data_type = "s"
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def apply_patches(proposals: list[PatchProposal], store: EvidenceStore, output_dir: Path) -> list[dict]:
    """Save new *_FIXED.xlsx copies after complete batch preflight and validation."""
    if not proposals:
        return []
    if len(proposals) > 100:
        raise ValueError("Patch batch exceeds the 100-cell bound")
    store.verify_originals()
    output_dir = Path(output_dir).resolve()
    grouped: dict[Path, list[PatchProposal]] = defaultdict(list)
    destinations: dict[Path, Path] = {}
    seen_cells = set()
    original_paths = {Path(item.document.original_storage_key).resolve() for item in store.documents}
    for proposal in proposals:
        source = _validate_proposal(proposal, store)
        target = output_dir / f"{source.stem}_FIXED.xlsx"
        identity = (source, proposal.sheet, proposal.cell)
        if identity in seen_cells:
            raise ValueError("Multiple patch proposals target the same cell")
        seen_cells.add(identity)
        if target in original_paths or os.path.lexists(target):
            raise ValueError("Corrected output would overwrite a source or existing path")
        if target in destinations and destinations[target] != source:
            raise ValueError("Different source workbooks collide at the same corrected filename")
        destinations[target] = source
        grouped[source].append(proposal)

    timestamp = datetime.now(timezone.utc).isoformat()
    prepared = []
    workbooks = []
    published: list[tuple[Path, int]] = []
    try:
        # Load and validate every source/target before creating any output file.
        with warnings.catch_warnings():
            warnings.simplefilter("error")
            for source, items in grouped.items():
                workbook = load_workbook(source, data_only=False, keep_links=True)
                workbooks.append(workbook)
                for proposal in items:
                    if proposal.sheet not in workbook.sheetnames:
                        raise ValueError("Patch worksheet is missing from the original workbook")
                    sheet = workbook[proposal.sheet]
                    cell = sheet[proposal.cell]
                    if (cell.data_type in ("f", "b", "e") or cell.value is None
                            or str(cell.value) != proposal.old_value
                            or any(proposal.cell in area for area in sheet.merged_cells.ranges)):
                        raise ValueError("Patch cell is a formula, merged, changed or ambiguous")
                    cell.value = _excel_number(proposal.new_value)
                _audit_trail(workbook, items, timestamp)
                prepared.append((source, workbook, items, output_dir / f"{source.stem}_FIXED.xlsx"))
            store.verify_originals()
            output_dir.mkdir(parents=True, exist_ok=True)
            with tempfile.TemporaryDirectory(prefix=".verified-patches-", dir=output_dir) as staging:
                staged, metadata = [], []
                for index, (source, workbook, items, target) in enumerate(prepared):
                    temporary = Path(staging) / f"{index}.xlsx"
                    workbook.save(temporary)
                    saved = load_workbook(temporary, data_only=False, keep_links=True)
                    original = load_workbook(source, data_only=False, keep_links=True)
                    try:
                        _assert_preserved(original, saved, items)
                    finally:
                        saved.close()
                        original.close()
                    source_hash = store.docs[items[0].document_id].document.document_hash
                    metadata.append({
                        "source_file": items[0].source_file, "source_path": str(source),
                        "output_file": str(target), "source_hash": source_hash,
                        "source_hash_after": _file_hash(source), "output_hash": _file_hash(temporary),
                        "patch_count": len(items), "patches": [item.model_dump(mode="json") for item in items],
                    })
                    staged.append((temporary, target))
                store.verify_originals()
                for temporary, target in staged:
                    # link is atomic and fails if the destination already exists;
                    # os.replace/rename could clobber a concurrent file instead.
                    os.link(temporary, target)
                    published.append((target, target.stat().st_ino))
                store.verify_originals()
                return metadata
    except Exception:
        for target, inode in reversed(published):
            if target.exists() and target.stat().st_ino == inode:
                target.unlink()
        raise
    finally:
        for workbook in workbooks:
            workbook.close()
