"""Bounded, non-executing normalization for text PDFs, XLSX, and CSV."""

from __future__ import annotations

import csv
import io
import mimetypes
import re
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pypdf import PdfReader

from .ids import sha256_bytes, stable_id
from .models import (
    DocumentRole,
    EvidenceKind,
    ExtractionStatus,
    NormalizedDocument,
    SourceDocument,
    SourceRef,
    WorkbookSheet,
)

MAX_FILE_BYTES = 25 * 1024 * 1024
MAX_PDF_PAGES = 250
MAX_PDF_PAGE_CHARS = 100_000
MAX_PDF_TOTAL_CHARS = 2_000_000
MAX_XLSX_UNCOMPRESSED_BYTES = 100 * 1024 * 1024
MAX_XLSX_MEMBERS = 10_000
MAX_WORKBOOK_ROWS = 20_000
MAX_WORKBOOK_COLUMNS = 500
MAX_NONEMPTY_CELLS = 100_000
MAX_CSV_ROWS = 50_000
MAX_CSV_COLUMNS = 500
MAX_CELL_CHARS = 32_000


class IngestionError(ValueError):
    def __init__(self, code: str, message: str) -> None:
        super().__init__(message)
        self.code = code
        self.message = message


@dataclass(frozen=True)
class RoleGuess:
    role: DocumentRole
    confident: bool
    rationale: str


def detect_document_role(filename: str) -> RoleGuess:
    name = Path(filename).name.lower()
    if "side" in name and "letter" in name:
        return RoleGuess(DocumentRole.SIDE_LETTER, True, "filename identifies a side letter")
    if "lpa" in name or "partnership_agreement" in name or "partnership agreement" in name:
        return RoleGuess(DocumentRole.LPA, True, "filename identifies an LPA")
    if "nav" in name and name.endswith(".xlsx"):
        return RoleGuess(DocumentRole.NAV_WORKBOOK, True, "filename identifies a NAV workbook")
    if "register" in name or "investor_input" in name or "investor input" in name:
        return RoleGuess(
            DocumentRole.INVESTOR_REGISTER,
            True,
            "filename identifies an investor/input register",
        )
    return RoleGuess(
        DocumentRole.SUPPORTING,
        False,
        "role could not be established from the filename; reviewer confirmation is required",
    )


def normalize_file(
    path: Path,
    role: DocumentRole | None = None,
    *,
    original_storage_key: str | None = None,
) -> NormalizedDocument:
    path = Path(path)
    if not path.is_file():
        raise IngestionError("FILE_NOT_FOUND", f"Source file does not exist: {path.name}")
    data = path.read_bytes()
    if len(data) > MAX_FILE_BYTES:
        raise IngestionError(
            "FILE_TOO_LARGE",
            f"{path.name} exceeds the {MAX_FILE_BYTES // (1024 * 1024)} MiB upload limit",
        )
    if not data:
        raise IngestionError("EMPTY_FILE", f"{path.name} is empty")

    guessed = detect_document_role(path.name)
    assigned_role = role or guessed.role
    document_hash = sha256_bytes(data)
    document_id = stable_id("doc", document_hash, path.name)
    storage_key = original_storage_key or f"sources/{document_id}/{path.name}"
    suffix = path.suffix.lower()

    if suffix == ".pdf":
        return _normalize_pdf(
            path,
            data,
            document_id,
            document_hash,
            assigned_role,
            storage_key,
            guessed.confident or role is not None,
        )
    if suffix == ".xlsx":
        return _normalize_xlsx(
            path,
            data,
            document_id,
            document_hash,
            assigned_role,
            storage_key,
            guessed.confident or role is not None,
        )
    if suffix == ".csv":
        return _normalize_csv(
            path,
            data,
            document_id,
            document_hash,
            assigned_role,
            storage_key,
            guessed.confident or role is not None,
        )
    raise IngestionError(
        "UNSUPPORTED_FORMAT",
        f"Unsupported source format for {path.name}; supported formats are text PDF, XLSX, and CSV",
    )


def _source_document(
    *,
    path: Path,
    data: bytes,
    document_id: str,
    document_hash: str,
    role: DocumentRole,
    storage_key: str,
    status: ExtractionStatus,
    warnings: list[str],
) -> SourceDocument:
    mime = {
        ".pdf": "application/pdf",
        ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ".csv": "text/csv",
    }.get(path.suffix.lower()) or mimetypes.guess_type(path.name)[0] or "application/octet-stream"
    return SourceDocument(
        document_id=document_id,
        filename=path.name,
        document_hash=document_hash,
        role=role,
        mime_type=mime,
        size_bytes=len(data),
        extraction_status=status,
        warnings=warnings,
        original_storage_key=storage_key,
    )


def _normalize_pdf(
    path: Path,
    data: bytes,
    document_id: str,
    document_hash: str,
    role: DocumentRole,
    storage_key: str,
    role_confirmed: bool,
) -> NormalizedDocument:
    if not data.startswith(b"%PDF-"):
        raise IngestionError("INVALID_PDF", f"{path.name} does not have a valid PDF signature")
    try:
        reader = PdfReader(io.BytesIO(data), strict=True)
    except Exception as exc:  # pypdf exposes several parser-specific exception types
        raise IngestionError("PDF_PARSE_FAILED", f"Could not parse {path.name}: {exc}") from exc
    if reader.is_encrypted:
        raise IngestionError("ENCRYPTED_PDF", f"Encrypted PDF is not supported: {path.name}")
    if len(reader.pages) > MAX_PDF_PAGES:
        raise IngestionError(
            "PDF_PAGE_LIMIT",
            f"{path.name} has {len(reader.pages)} pages; limit is {MAX_PDF_PAGES}",
        )

    warnings: list[str] = []
    evidence: list[SourceRef] = []
    total_chars = 0
    section: str | None = None
    for page_number, page in enumerate(reader.pages, start=1):
        try:
            text = page.extract_text() or ""
        except Exception as exc:
            warnings.append(f"Page {page_number} could not be extracted: {type(exc).__name__}")
            continue
        if len(text) > MAX_PDF_PAGE_CHARS:
            warnings.append(f"Page {page_number} text was truncated at {MAX_PDF_PAGE_CHARS} characters")
            text = text[:MAX_PDF_PAGE_CHARS]
        total_chars += len(text)
        if total_chars > MAX_PDF_TOTAL_CHARS:
            raise IngestionError(
                "PDF_TEXT_LIMIT",
                f"Extracted text in {path.name} exceeds {MAX_PDF_TOTAL_CHARS} characters",
            )
        offset = 0
        for raw_block in re.split(r"\n\s*\n|(?<=\.)\s*\n", text):
            block = " ".join(raw_block.split()).strip()
            if not block:
                continue
            located = text.find(raw_block, offset)
            start = located if located >= 0 else offset
            end = start + len(raw_block)
            offset = max(end, offset)
            if re.match(r"^(?:section|article|schedule|\d+(?:\.\d+)*)\b", block, re.I):
                section = block[:160]
            evidence.append(
                SourceRef(
                    evidence_id=stable_id(
                        "ev",
                        document_hash,
                        "pdf",
                        page_number,
                        start,
                        end,
                        block,
                    ),
                    document_id=document_id,
                    document_hash=document_hash,
                    kind=EvidenceKind.PDF_TEXT,
                    page=page_number,
                    section=section,
                    text_start=start,
                    text_end=end,
                    quote=block,
                )
            )

    if not evidence:
        raise IngestionError(
            "IMAGE_ONLY_OR_EMPTY_PDF",
            f"No usable text was extracted from {path.name}; image-only PDFs require OCR",
        )
    if not role_confirmed:
        warnings.append("Document role requires reviewer confirmation")
    status = ExtractionStatus.PARTIAL if warnings else ExtractionStatus.COMPLETE
    if not role_confirmed:
        status = ExtractionStatus.NEEDS_CONFIRMATION
    return NormalizedDocument(
        document=_source_document(
            path=path,
            data=data,
            document_id=document_id,
            document_hash=document_hash,
            role=role,
            storage_key=storage_key,
            status=status,
            warnings=warnings,
        ),
        evidence=evidence,
        layout={"page_count": len(reader.pages), "text_based": True},
    )


def _inspect_xlsx_zip(path: Path, data: bytes) -> None:
    if not zipfile.is_zipfile(io.BytesIO(data)):
        raise IngestionError("INVALID_XLSX", f"{path.name} is not a valid XLSX package")
    with zipfile.ZipFile(io.BytesIO(data)) as archive:
        members = archive.infolist()
        if len(members) > MAX_XLSX_MEMBERS:
            raise IngestionError(
                "XLSX_MEMBER_LIMIT",
                f"{path.name} contains too many package members",
            )
        decompressed = sum(member.file_size for member in members)
        if decompressed > MAX_XLSX_UNCOMPRESSED_BYTES:
            raise IngestionError(
                "XLSX_DECOMPRESSED_LIMIT",
                f"{path.name} exceeds the {MAX_XLSX_UNCOMPRESSED_BYTES // (1024 * 1024)} MiB decompressed limit",
            )
        unsafe = [
            member.filename
            for member in members
            if member.filename.startswith("/") or ".." in Path(member.filename).parts
        ]
        if unsafe:
            raise IngestionError("UNSAFE_XLSX_PACKAGE", f"Unsafe XLSX member path in {path.name}")


def _normalize_xlsx(
    path: Path,
    data: bytes,
    document_id: str,
    document_hash: str,
    role: DocumentRole,
    storage_key: str,
    role_confirmed: bool,
) -> NormalizedDocument:
    _inspect_xlsx_zip(path, data)
    try:
        formulas_book = load_workbook(
            io.BytesIO(data),
            read_only=False,
            data_only=False,
            keep_links=False,
        )
        values_book = load_workbook(
            io.BytesIO(data),
            read_only=False,
            data_only=True,
            keep_links=False,
        )
    except Exception as exc:
        raise IngestionError("XLSX_PARSE_FAILED", f"Could not parse {path.name}: {exc}") from exc

    warnings: list[str] = []
    evidence: list[SourceRef] = []
    sheets: list[WorkbookSheet] = []
    nonempty_cells = 0
    try:
        for sheet in formulas_book.worksheets:
            if sheet.max_row > MAX_WORKBOOK_ROWS or sheet.max_column > MAX_WORKBOOK_COLUMNS:
                raise IngestionError(
                    "WORKSHEET_DIMENSION_LIMIT",
                    f"Sheet {sheet.title!r} exceeds the {MAX_WORKBOOK_ROWS}x{MAX_WORKBOOK_COLUMNS} limit",
                )
            cached_sheet = values_book[sheet.title]
            sheets.append(
                WorkbookSheet(
                    name=sheet.title,
                    max_row=sheet.max_row,
                    max_column=sheet.max_column,
                    hidden=sheet.sheet_state != "visible",
                    frozen_panes=str(sheet.freeze_panes) if sheet.freeze_panes else None,
                    merged_ranges=[str(item) for item in sheet.merged_cells.ranges][:500],
                )
            )
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                    nonempty_cells += 1
                    if nonempty_cells > MAX_NONEMPTY_CELLS:
                        raise IngestionError(
                            "WORKBOOK_CELL_LIMIT",
                            f"{path.name} exceeds the {MAX_NONEMPTY_CELLS} non-empty cell limit",
                        )
                    original = str(cell.value)
                    if len(original) > MAX_CELL_CHARS:
                        warnings.append(f"{sheet.title}!{cell.coordinate} was truncated")
                        original = original[:MAX_CELL_CHARS]
                    formula = original if cell.data_type == "f" else None
                    cached = cached_sheet[cell.coordinate].value if formula else None
                    cache_status = (
                        "PRESENT_UNVERIFIED"
                        if formula and cached is not None
                        else "MISSING"
                        if formula
                        else "NOT_APPLICABLE"
                    )
                    if formula and cached is None:
                        warnings.append(
                            f"{sheet.title}!{cell.coordinate} contains a formula without a usable cached value"
                        )
                    evidence.append(
                        SourceRef(
                            evidence_id=stable_id(
                                "ev",
                                document_hash,
                                "xlsx",
                                sheet.title,
                                cell.coordinate,
                                original,
                            ),
                            document_id=document_id,
                            document_hash=document_hash,
                            kind=EvidenceKind.WORKBOOK_CELL,
                            sheet=sheet.title,
                            cell=cell.coordinate,
                            original_value=original,
                            normalized_value=str(cached) if cached is not None else original,
                            formula=formula,
                            cached_value=str(cached) if cached is not None else None,
                            cache_status=cache_status,
                            data_type=str(cell.data_type),
                            number_format=cell.number_format,
                        )
                    )
    finally:
        formulas_book.close()
        values_book.close()

    if not role_confirmed:
        warnings.append("Document role requires reviewer confirmation")
    status = ExtractionStatus.PARTIAL if warnings else ExtractionStatus.COMPLETE
    if not role_confirmed:
        status = ExtractionStatus.NEEDS_CONFIRMATION
    return NormalizedDocument(
        document=_source_document(
            path=path,
            data=data,
            document_id=document_id,
            document_hash=document_hash,
            role=role,
            storage_key=storage_key,
            status=status,
            warnings=warnings,
        ),
        evidence=evidence,
        workbook_sheets=sheets,
        layout={"sheet_count": len(sheets), "nonempty_cells": nonempty_cells},
    )


def _decode_csv(path: Path, data: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise IngestionError("CSV_ENCODING", f"{path.name} must use UTF-8 encoding")


def _normalize_csv(
    path: Path,
    data: bytes,
    document_id: str,
    document_hash: str,
    role: DocumentRole,
    storage_key: str,
    role_confirmed: bool,
) -> NormalizedDocument:
    text = _decode_csv(path, data)
    try:
        dialect = csv.Sniffer().sniff(text[:8192], delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    reader = csv.reader(io.StringIO(text, newline=""), dialect)
    try:
        raw_headers = next(reader)
    except StopIteration as exc:
        raise IngestionError("EMPTY_CSV", f"{path.name} has no header row") from exc
    if len(raw_headers) > MAX_CSV_COLUMNS:
        raise IngestionError("CSV_COLUMN_LIMIT", f"{path.name} has too many columns")
    headers = _normalise_headers(raw_headers)
    warnings: list[str] = []
    evidence: list[SourceRef] = []
    row_count = 1
    for row_number, row in enumerate(reader, start=2):
        row_count = row_number
        if row_number > MAX_CSV_ROWS + 1:
            raise IngestionError("CSV_ROW_LIMIT", f"{path.name} exceeds {MAX_CSV_ROWS} data rows")
        if len(row) > MAX_CSV_COLUMNS:
            raise IngestionError("CSV_COLUMN_LIMIT", f"Row {row_number} has too many columns")
        padded = row + [""] * (len(headers) - len(row))
        for index, value in enumerate(padded[: len(headers)]):
            if value == "":
                continue
            if len(value) > MAX_CELL_CHARS:
                warnings.append(f"Row {row_number}, {headers[index]} was truncated")
                value = value[:MAX_CELL_CHARS]
            evidence.append(
                SourceRef(
                    evidence_id=stable_id(
                        "ev",
                        document_hash,
                        "csv",
                        row_number,
                        headers[index],
                        value,
                    ),
                    document_id=document_id,
                    document_hash=document_hash,
                    kind=EvidenceKind.CSV_CELL,
                    csv_row=row_number,
                    csv_column=headers[index],
                    original_value=value,
                    normalized_value=value.strip(),
                    data_type="string",
                )
            )
    if not role_confirmed:
        warnings.append("Document role requires reviewer confirmation")
    status = ExtractionStatus.PARTIAL if warnings else ExtractionStatus.COMPLETE
    if not role_confirmed:
        status = ExtractionStatus.NEEDS_CONFIRMATION
    return NormalizedDocument(
        document=_source_document(
            path=path,
            data=data,
            document_id=document_id,
            document_hash=document_hash,
            role=role,
            storage_key=storage_key,
            status=status,
            warnings=warnings,
        ),
        evidence=evidence,
        csv_headers=headers,
        layout={"row_count": row_count, "column_count": len(headers)},
    )


def _normalise_headers(raw_headers: Iterable[str]) -> list[str]:
    headers: list[str] = []
    seen: dict[str, int] = {}
    for index, raw in enumerate(raw_headers, start=1):
        base = raw.strip() or f"column_{get_column_letter(index)}"
        count = seen.get(base, 0) + 1
        seen[base] = count
        headers.append(base if count == 1 else f"{base}_{count}")
    return headers
