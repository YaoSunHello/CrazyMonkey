"""Generate the fictional source pack used by the offline synthetic demo."""

from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib.colors import HexColor
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas


FUND_NAME = "Example Growth Fund III"
REPORTING_PERIOD = "Q3 2026"


def generate_synthetic_pack(output_dir: Path) -> dict[str, object]:
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    lpa_path = output_dir / "Example_Growth_Fund_III_LPA.pdf"
    _write_pdf(
        lpa_path,
        title="Example Growth Fund III",
        subtitle="Limited Partnership Agreement — fictional synthetic demo",
        sections=[
            (
                "Section 1 — Scope and period",
                [
                    "This fictional agreement is created solely for the CrazyMonkey synthetic demonstration.",
                    "The review period is Q3 2026, from 1 July 2026 through 30 September 2026.",
                ],
            ),
            (
                "Section 8.1 — Management fee",
                [
                    "The default annual management fee is 2.0% of the applicable investor Fee Base.",
                    "For Q3 2026 the quarterly fee is annual rate x 0.25 x Fee Base.",
                    "Management fees are denominated in GBP, with no other adjustments for this synthetic case.",
                    "The calculated fee is rounded to the nearest penny using round-half-up; comparison tolerance is GBP 0.01.",
                ],
            ),
            (
                "Section 8.2 — Investor-specific terms",
                [
                    "A side-letter term applies only where investor identity, management-fee scope, effective date, and governing relationship are supported by the supplied evidence.",
                    "A future or expired override does not replace the default rate for the review period.",
                ],
            ),
        ],
    )

    side_letters = {
        "LP01": {
            "rate": None,
            "effective": "1 January 2025",
            "text": "No management-fee variation is granted; the LPA default remains applicable.",
        },
        "LP02": {
            "rate": "1.5%",
            "effective": "1 January 2025",
            "text": "The annual management fee applicable to LP02 is 1.5% of the Fee Base.",
        },
        "LP03": {
            "rate": "1.5%",
            "effective": "1 January 2026",
            "text": "The annual management fee applicable to LP03 is 1.5% of the Fee Base.",
        },
        "LP04": {
            "rate": None,
            "effective": "1 January 2025",
            "text": "No management-fee rate variation is granted; the LPA default remains applicable.",
        },
        "LP05": {
            "rate": "1.5%",
            "effective": "1 October 2026",
            "text": "The annual management fee applicable to LP05 is 1.5% of the Fee Base.",
        },
    }
    side_letter_paths: list[Path] = []
    for investor_id, term in side_letters.items():
        path = output_dir / f"{investor_id}_Side_Letter.pdf"
        _write_pdf(
            path,
            title=f"{investor_id} Side Letter",
            subtitle="Fictional synthetic investor agreement",
            sections=[
                (
                    "Section 1 — Investor identity",
                    [
                        f"Investor ID: {investor_id}.",
                        f"This letter supplements the Example Growth Fund III LPA for {investor_id} only.",
                    ],
                ),
                (
                    "Section 3.1 — Management fee term",
                    [
                        str(term["text"]),
                        f"Effective from {term['effective']}; no end date is specified.",
                    ],
                ),
            ],
        )
        side_letter_paths.append(path)

    register_path = output_dir / "investor_input_register.csv"
    register_rows = [
        ["LP01", "Northbridge Pension", "10000000", "GBP", "YES", "LP01_Side_Letter.pdf"],
        ["LP02", "Meridian Endowment", "10000000", "GBP", "YES", "LP02_Side_Letter.pdf"],
        ["LP03", "Cedar Grove Foundation", "10000000", "GBP", "YES", "LP03_Side_Letter.pdf"],
        ["LP04", "Harbour Family Office", "8000000", "GBP", "YES", "LP04_Side_Letter.pdf"],
        ["LP05", "Sterling University", "10000000", "GBP", "YES", "LP05_Side_Letter.pdf"],
        ["LP06", "Westgate Charitable Trust", "10000000", "GBP", "YES", "LP06_Side_Letter.pdf"],
    ]
    with register_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(
            [
                "investor_id",
                "investor_name",
                "fee_base",
                "currency",
                "side_letter_expected",
                "side_letter_filename",
            ]
        )
        writer.writerows(register_rows)

    nav_path = output_dir / "Administrator_NAV_Q3_2026.xlsx"
    _write_administrator_workbook(nav_path)

    manifest = {
        "fixture": "CrazyMonkey fictional synthetic management-fee pack",
        "fund_name": FUND_NAME,
        "reporting_period": REPORTING_PERIOD,
        "synthetic": True,
        "files": [
            {"filename": lpa_path.name, "role": "LPA"},
            {"filename": nav_path.name, "role": "NAV_WORKBOOK"},
            {"filename": register_path.name, "role": "INVESTOR_REGISTER"},
            *[
                {"filename": item.name, "role": "SIDE_LETTER"}
                for item in sorted(side_letter_paths)
            ],
        ],
        "notice": "All people, entities, terms, documents, and amounts are fictional synthetic data.",
    }
    manifest_path = output_dir / "manifest.json"
    manifest_path.write_text(json.dumps(manifest, indent=2) + "\n", encoding="utf-8")
    return manifest


def _write_pdf(
    path: Path,
    *,
    title: str,
    subtitle: str,
    sections: list[tuple[str, list[str]]],
) -> None:
    page = canvas.Canvas(str(path), pagesize=A4, invariant=1, pageCompression=1)
    width, height = A4
    page.setTitle(title)
    page.setAuthor("CrazyMonkey synthetic fixture generator")
    page.setFillColor(HexColor("#143C46"))
    page.rect(0, height - 92, width, 92, stroke=0, fill=1)
    page.setFillColor(HexColor("#FFFFFF"))
    page.setFont("Helvetica-Bold", 20)
    page.drawString(44, height - 50, title)
    page.setFont("Helvetica", 10)
    page.drawString(44, height - 70, subtitle)

    y = height - 126
    for heading, paragraphs in sections:
        page.setFillColor(HexColor("#143C46"))
        page.setFont("Helvetica-Bold", 12)
        page.drawString(44, y, heading)
        y -= 22
        page.setFillColor(HexColor("#263238"))
        page.setFont("Helvetica", 10)
        for paragraph in paragraphs:
            for line in _wrap(paragraph, 92):
                page.drawString(52, y, line)
                y -= 15
            y -= 5
        y -= 8
    page.setFillColor(HexColor("#6A7377"))
    page.setFont("Helvetica-Oblique", 8)
    page.drawString(
        44,
        30,
        "Synthetic demo document — not legal advice and not a real fund agreement.",
    )
    page.save()


def _wrap(text: str, width: int) -> list[str]:
    words = text.split()
    lines: list[str] = []
    current: list[str] = []
    for word in words:
        if current and len(" ".join([*current, word])) > width:
            lines.append(" ".join(current))
            current = [word]
        else:
            current.append(word)
    if current:
        lines.append(" ".join(current))
    return lines


def _write_administrator_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Investor Fees"
    sheet["A1"] = "Example Growth Fund III — Q3 2026 NAV"
    sheet["A2"] = "Synthetic administrator return — management fees are hard-coded values"
    headers = [
        "Investor ID",
        "Investor Name",
        "Fee Base Used",
        "Annual Rate Used",
        "Period Factor",
        "Reported Fee",
        "Currency",
        "Value Provenance",
    ]
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(row=3, column=column, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1E5966")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    rows = [
        ["LP01", "Northbridge Pension", 10_000_000, 0.0200, 0.25, 50_000, "GBP", "Hard-coded by administrator"],
        ["LP02", "Meridian Endowment", 10_000_000, 0.0150, 0.25, 37_500, "GBP", "Hard-coded by administrator"],
        ["LP03", "Cedar Grove Foundation", 10_000_000, 0.0200, 0.25, 50_000, "GBP", "Hard-coded by administrator"],
        ["LP04", "Harbour Family Office", 10_000_000, 0.0200, 0.25, 50_000, "GBP", "Hard-coded by administrator"],
        ["LP05", "Sterling University", 10_000_000, 0.0200, 0.25, 50_000, "GBP", "Hard-coded by administrator"],
        ["LP06", "Westgate Charitable Trust", 10_000_000, 0.0150, 0.25, 37_500, "GBP", "Hard-coded by administrator"],
    ]
    for row_index, row in enumerate(rows, start=4):
        for column_index, value in enumerate(row, start=1):
            sheet.cell(row=row_index, column=column_index, value=value)
    for row in range(4, 10):
        sheet.cell(row=row, column=3).number_format = '£#,##0.00'
        sheet.cell(row=row, column=4).number_format = '0.00%'
        sheet.cell(row=row, column=5).number_format = '0.00'
        sheet.cell(row=row, column=6).number_format = '£#,##0.00'
    widths = [14, 28, 18, 18, 14, 18, 12, 29]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = "A4"
    sheet.auto_filter.ref = "A3:H9"
    sheet.sheet_view.showGridLines = False
    workbook.properties.title = "CrazyMonkey synthetic administrator NAV workbook"
    workbook.properties.creator = "CrazyMonkey synthetic fixture generator"
    workbook.save(path)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("fixtures/synthetic_q3_2026"),
        help="Directory for generated synthetic source documents",
    )
    args = parser.parse_args()
    manifest = generate_synthetic_pack(args.output)
    print(json.dumps(manifest, indent=2))


if __name__ == "__main__":
    main()
