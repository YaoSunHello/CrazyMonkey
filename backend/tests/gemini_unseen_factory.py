"""Create one independent synthetic source pair for blind live-model validation.

The command emits only a neutral input path. The evaluator's answer key and
source hashes are siblings of that input directory and are never model inputs.
All organizations, account terms and figures in this test factory are fictional.
"""

from __future__ import annotations

import argparse
import hashlib
import json
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate


def _write_statement(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Remuneration statement"
    sheet.sheet_view.showGridLines = False
    for column, width in {"A": 3, "B": 3, "C": 37, "D": 3, "E": 48}.items():
        sheet.column_dimensions[column].width = width
    sheet["C3"] = "Half-year remuneration statement"
    sheet["C3"].font = Font(name="Arial", size=16, bold=True, color="243447")
    sheet["C4"] = "Fictional administrator statement. All monetary amounts in USD."
    sheet["C4"].font = Font(name="Arial", size=11, italic=True, color="596572")
    sheet["C7"] = "Statement field"
    sheet["E7"] = "Administrator record"
    for row in sheet.iter_rows(min_row=7, max_row=7, min_col=3, max_col=5):
        for cell in row:
            cell.fill = PatternFill("solid", fgColor="34495E")
            cell.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    records = [
        ("Fund legal name", "Cedar Estuary Opportunities III"),
        ("Investor account reference", "IM-927"),
        ("Investor legal name", "Harbour Mathematics Foundation"),
        ("Accounting period starts", "2028-01-01"),
        ("Accounting period ends", "2028-06-30"),
        ("Account currency", "USD"),
        ("Charge category", "Stewardship remuneration (management fee)"),
        ("Invested-capital fee base", 6840000),
        ("Semiannual allocation factor", 0.5),
        ("Stewardship remuneration booked", 54720),
        ("Governing agreement", "terms_64.pdf"),
    ]
    for row, (label, value) in enumerate(records, start=8):
        sheet.cell(row, 3, label)
        sheet.cell(row, 5, value)
        sheet.row_dimensions[row].height = 29
        for column in (3, 5):
            sheet.cell(row, column).font = Font(name="Arial", size=11, color="243447")
            sheet.cell(row, column).alignment = Alignment(vertical="center", wrap_text=True)
    for coordinate, number_format in {
        "E15": '"USD "#,##0.00', "E16": "0.00", "E17": '"USD "#,##0.00',
    }.items():
        sheet[coordinate].number_format = number_format
    for cell in sheet[17][2:5]:
        cell.border = Border(top=Side(style="thin", color="ADB9C5"))
        cell.font = Font(name="Arial", size=11, bold=True, color="243447")
    sheet["C21"] = "The booked amount is the administrator's charge for this full period."
    sheet["C22"] = "No payments, adjustments or withholding entries are included."
    for coordinate in ("C21", "C22"):
        sheet[coordinate].font = Font(name="Arial", size=10, italic=True, color="596572")
    sheet.print_options.horizontalCentered = True
    sheet.print_area = "C3:E22"
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 1
    workbook.save(path)
    workbook.close()


def _write_terms(path: Path) -> None:
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle("TermBody", fontName="Helvetica", fontSize=11,
                              leading=15, spaceAfter=11))
    styles.add(ParagraphStyle("TermHeading", fontName="Helvetica-Bold", fontSize=12,
                              leading=16, spaceBefore=9, spaceAfter=7,
                              textColor=colors.HexColor("#34495E")))
    paragraphs = [
        ("Title", "Investor remuneration agreement"),
        ("TermBody", "Cedar Estuary Opportunities III. Fictional source document."),
        ("TermHeading", "1. Parties, authority and period"),
        ("TermBody", "This executed agreement is between Cedar Estuary Opportunities III "
         "and Harbour Mathematics Foundation, investor account IM-927. It was executed "
         "on 18 December 2027 and takes effect on 1 January 2028. It governs the full "
         "accounting period from 1 January 2028 through 30 June 2028, inclusive."),
        ("TermBody", "This is the sole operative agreement for this investor's "
         "management fee in that fund and period. It supersedes all earlier fee terms. "
         "There are no side letters, concessions or subsequent amendments affecting "
         "that period."),
        ("TermHeading", "2. Charge definition and contractual annual rate"),
        ("TermBody", "Stewardship remuneration means the recurring management fee "
         "payable by investor IM-927 to Cedar Estuary Opportunities III. The applicable "
         "annual rate for the entire accounting period is 1.35%. All charges and "
         "the fee base are denominated in United States dollars (USD)."),
        ("TermHeading", "3. Agreed invested-capital fee base"),
        ("TermBody", "For Harbour Mathematics Foundation, account IM-927, the agreed "
         "invested-capital fee base is USD 6,840,000.00 throughout 1 January 2028 "
         "through 30 June 2028. This fixed amount is the exact management-fee base. "
         "No capital movements or changes in fee base occur during this period."),
        ("TermHeading", "4. Period allocation and calculation"),
        ("TermBody", "The accounting period is one full half-year. Its contractual "
         "semiannual allocation factor is exactly 0.50. Calculate this period's "
         "stewardship remuneration by multiplying the invested-capital fee base by "
         "the annual rate expressed as a decimal and by the semiannual allocation "
         "factor. Actual-day prorating is not used."),
        ("TermHeading", "5. Rounding and reconciliation"),
        ("TermBody", "Round the resulting charge to two decimal places using round "
         "half up. Reconcile it against the administrator's booked stewardship "
         "remuneration for the same investor, fund, currency and accounting period. "
         "The permitted reconciliation tolerance is USD 0.01. There are no minimum "
         "charges, caps, offsets, taxes, rebates, withholding or other adjustments."),
    ]
    story = [Paragraph(text, styles[style]) for style, text in paragraphs]

    def footer(canvas, document) -> None:
        canvas.setFont("Helvetica", 8)
        canvas.setFillColor(colors.HexColor("#596572"))
        canvas.drawString(18 * mm, 14 * mm, "Fictional source pack - no real fund or investor")
        canvas.drawRightString(A4[0] - 18 * mm, 14 * mm, f"Page {document.page}")

    SimpleDocTemplate(str(path), pagesize=A4, leftMargin=18 * mm,
                      rightMargin=18 * mm, topMargin=17 * mm, bottomMargin=23 * mm,
                      title="Investor remuneration agreement", author="Independent QA",
                      invariant=1).build(story, onFirstPage=footer, onLaterPages=footer)


def generate_gemini_blind_pair(output_dir: Path) -> dict:
    output_dir = output_dir.resolve()
    inputs, control = output_dir / "input", output_dir / "control"
    # Refuse any pre-existing root to keep the first blind trial immutable.
    output_dir.mkdir(parents=True, exist_ok=False)
    inputs.mkdir()
    control.mkdir()
    _write_statement(inputs / "statement_93.xlsx")
    _write_terms(inputs / "terms_64.pdf")
    fee_base = Decimal("6840000.00")
    rate = Decimal("0.0135")
    factor = Decimal("0.50")
    expected = (fee_base * rate * factor).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    reported = Decimal("54720.00")
    oracle = {
        "synthetic": True,
        "investor_id": "IM-927",
        "investor_name": "Harbour Mathematics Foundation",
        "fund_name": "Cedar Estuary Opportunities III",
        "period_start": "2028-01-01",
        "period_end": "2028-06-30",
        "status": "DISCREPANCY",
        "currency": "USD",
        "fee_base": str(fee_base),
        "annual_rate": str(rate),
        "period_factor": str(factor),
        "expected_amount": str(expected),
        "reported_amount": str(reported),
        "difference_reported_minus_expected": str(reported - expected),
        "reason": "The booked stewardship remuneration exceeds the contractual half-year management fee.",
        "source_anchors": {
            "reported": "statement_93.xlsx / Remuneration statement!E17",
            "base": "statement_93.xlsx / Remuneration statement!E15; terms_64.pdf page 1 section 3",
            "factor": "statement_93.xlsx / Remuneration statement!E16; terms_64.pdf page 1 section 4",
            "annual_rate": "terms_64.pdf page 1 section 2",
            "rounding_tolerance": "terms_64.pdf page 1 section 5",
            "identity_period_authority": "terms_64.pdf page 1 section 1",
        },
    }
    hashes = {path.name: hashlib.sha256(path.read_bytes()).hexdigest()
              for path in sorted(inputs.iterdir())}
    (control / "oracle.json").write_text(json.dumps(oracle, indent=2) + "\n")
    (control / "input_hashes.json").write_text(json.dumps(hashes, indent=2) + "\n")
    return {"input_dir": str(inputs)}


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()
    print(json.dumps(generate_gemini_blind_pair(args.output)))


if __name__ == "__main__":
    main()
