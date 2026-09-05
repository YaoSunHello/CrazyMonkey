"""Adversarial ingestion regressions using independent synthetic source files."""

from __future__ import annotations

import io
import tempfile
import unittest
import zipfile
from datetime import datetime
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.formula import ArrayFormula
from pypdf import PdfReader, PdfWriter
from reportlab.pdfgen import canvas

from app.atlas import ingestion
from app.atlas.ingestion import IngestionError, normalize_file
from app.atlas.models import DocumentRole, ExtractionStatus


class IngestionQATests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp = tempfile.TemporaryDirectory()
        self.addCleanup(self.temp.cleanup)
        self.root = Path(self.temp.name)

    def source(self, filename: str, content: str | bytes) -> Path:
        path = self.root / filename
        path.write_bytes(content.encode("utf-8") if isinstance(content, str) else content)
        return path

    def assert_ingestion_error(self, path: Path, code: str) -> None:
        with self.assertRaises(IngestionError) as caught:
            normalize_file(path, DocumentRole.SUPPORTING)
        self.assertEqual(caught.exception.code, code)

    def workbook(self, name: str = "unseen.xlsx") -> tuple[Workbook, Path]:
        book = Workbook()
        self.addCleanup(book.close)
        return book, self.root / name

    def test_csv_overwide_row_is_not_silently_discarded(self) -> None:
        path = self.source("wide.csv", "investor,fee\nJuniper,120,unmapped override\n")
        self.assert_ingestion_error(path, "CSV_ROW_WIDTH_MISMATCH")

    def test_filename_substrings_do_not_confirm_a_financial_document_role(self) -> None:
        for filename in ("wallpaper.pdf", "unavailable.xlsx", "discarded_side_letterhead.pdf", "unregistered.csv"):
            with self.subTest(filename=filename):
                guess = ingestion.detect_document_role(filename)
                self.assertEqual(guess.role, DocumentRole.SUPPORTING)
                self.assertFalse(guess.confident)
        for filename, role in (
            ("Example_Growth_Fund_III_LPA.pdf", DocumentRole.LPA),
            ("LP03_Side_Letter.pdf", DocumentRole.SIDE_LETTER),
            ("Administrator_NAV_Q3_2026.xlsx", DocumentRole.NAV_WORKBOOK),
            ("investor_input_register.csv", DocumentRole.INVESTOR_REGISTER),
        ):
            with self.subTest(filename=filename):
                guess = ingestion.detect_document_role(filename)
                self.assertEqual(guess.role, role)
                self.assertTrue(guess.confident)

    def test_duplicate_headers_do_not_collide_with_existing_suffixes(self) -> None:
        path = self.source("duplicate.csv", "fee,fee,fee_2,fee\n7,7,7,7\n")
        result = normalize_file(path, DocumentRole.SUPPORTING)
        self.assertEqual(len(set(result.csv_headers)), 4)
        self.assertEqual(len({item.evidence_id for item in result.evidence}), 4)
        self.assertEqual(len({item.locator for item in result.evidence}), 4)

    def test_byte_identical_named_documents_have_disjoint_evidence_ids(self) -> None:
        book, workbook_path = self.workbook()
        book.active["B2"] = "Cedar charge 120"
        book.save(workbook_path)
        pdf_buffer = io.BytesIO()
        pdf = canvas.Canvas(pdf_buffer)
        pdf.drawString(50, 700, "Cedar charge 120")
        pdf.save()
        for suffix, content in (
            ("csv", b"party,charge\nCedar,120\n"),
            ("xlsx", workbook_path.read_bytes()),
            ("pdf", pdf_buffer.getvalue()),
        ):
            with self.subTest(suffix=suffix):
                first_path = self.source(f"first.{suffix}", content)
                first = normalize_file(first_path, DocumentRole.SUPPORTING)
                second = normalize_file(self.source(f"second.{suffix}", content), DocumentRole.SUPPORTING)
                repeated = normalize_file(first_path, DocumentRole.SUPPORTING)
                self.assertNotEqual(first.document.document_id, second.document.document_id)
                self.assertEqual(first.document.document_hash, second.document.document_hash)
                self.assertTrue(
                    {item.evidence_id for item in first.evidence}.isdisjoint(
                        item.evidence_id for item in second.evidence
                    )
                )
                self.assertEqual(
                    [item.evidence_id for item in first.evidence],
                    [item.evidence_id for item in repeated.evidence],
                )

    def test_csv_unterminated_quote_is_a_typed_parse_error(self) -> None:
        self.assert_ingestion_error(
            self.source("broken.csv", 'investor,fee\n"Juniper,120\n'), "CSV_PARSE_FAILED"
        )

    def test_csv_parser_field_limit_is_a_typed_error_for_header_and_data(self) -> None:
        for name, contents in (
            ("header.csv", "x" * 200_000 + "\n1\n"),
            ("data.csv", "label\n" + "x" * 200_000 + "\n"),
        ):
            with self.subTest(name=name):
                self.assert_ingestion_error(self.source(name, contents), "CSV_PARSE_FAILED")

    def test_csv_header_length_is_bounded(self) -> None:
        with patch.object(ingestion, "MAX_CELL_CHARS", 8):
            self.assert_ingestion_error(self.source("header.csv", "long_header\n1\n"), "CSV_HEADER_LIMIT")

    def test_blank_and_header_only_csv_are_not_complete(self) -> None:
        for index, content in enumerate(("name,fee\n", "name,fee\n,\n", "name,fee\n  , \n", "\n")):
            with self.subTest(content=repr(content)):
                self.assert_ingestion_error(self.source(f"blank{index}.csv", content), "EMPTY_CSV")

    def test_short_csv_row_is_visibly_partial(self) -> None:
        result = normalize_file(self.source("short.csv", "name,fee\nJuniper\n"), DocumentRole.SUPPORTING)
        self.assertEqual(result.document.extraction_status, ExtractionStatus.PARTIAL)
        self.assertTrue(any("Row 2" in item for item in result.document.warnings))

    def test_csv_dialect_guess_does_not_discard_leading_value_spaces(self) -> None:
        path = self.source("spaced.csv", "party, fee\nCedar,  20\nJuniper,  30\n")
        result = normalize_file(path, DocumentRole.SUPPORTING)
        self.assertEqual(result.evidence[1].original_value, "  20")

    def test_csv_spaced_quoted_fields_obey_inferred_dialect(self) -> None:
        for quoted_amount, expected in (("50,000", "50,000"), ("50000", "50000"), ("50\n000", "50\n000")):
            with self.subTest(quoted_amount=quoted_amount):
                path = self.source("quoted.csv", f'investor, amount\n"LP, 03", "{quoted_amount}"\n')
                result = normalize_file(path, DocumentRole.SUPPORTING)
                self.assertEqual(len(result.evidence), 2)
                self.assertEqual(result.evidence[1].original_value, expected)
                self.assertEqual(result.document.extraction_status, ExtractionStatus.PARTIAL)
                self.assertTrue(any("separator whitespace" in item for item in result.document.warnings))

    def test_array_formula_fails_with_stable_explicit_unsupported_error(self) -> None:
        book, path = self.workbook()
        book.active["A1"] = ArrayFormula(ref="A1:A3", text="=ROW(A1:A3)")
        book.save(path)
        self.assert_ingestion_error(path, "UNSUPPORTED_XLSX_FORMULA")

    def test_excel_error_cells_require_review(self) -> None:
        book, path = self.workbook()
        book.active["B2"] = "#DIV/0!"
        book.save(path)
        result = normalize_file(path, DocumentRole.SUPPORTING)
        self.assertEqual(result.document.extraction_status, ExtractionStatus.PARTIAL)
        self.assertTrue(any("error" in item.lower() for item in result.document.warnings))
        self.assertEqual(result.evidence[0].original_value, "#DIV/0!")

    def test_blank_workbook_is_not_complete(self) -> None:
        for whitespace in (False, True):
            with self.subTest(whitespace=whitespace):
                book, path = self.workbook()
                if whitespace:
                    book.active["B3"] = "   "
                book.save(path)
                self.assert_ingestion_error(path, "EMPTY_XLSX")

    def test_pdf_blank_page_is_visibly_partial(self) -> None:
        page_buffer = io.BytesIO()
        writer = canvas.Canvas(page_buffer)
        writer.drawString(72, 700, "Section 1. Juniper annual charge is 1.25 percent.")
        writer.save()
        pdf = PdfWriter()
        pdf.add_page(PdfReader(io.BytesIO(page_buffer.getvalue())).pages[0])
        pdf.add_blank_page(width=612, height=792)
        target = io.BytesIO()
        pdf.write(target)
        result = normalize_file(self.source("unseen.pdf", target.getvalue()), DocumentRole.SUPPORTING)
        self.assertEqual(result.layout["page_count"], 2)
        self.assertEqual(result.document.extraction_status, ExtractionStatus.PARTIAL)
        self.assertTrue(any("Page 2" in item for item in result.document.warnings))

    def test_pdf_lazy_page_tree_parse_error_is_typed(self) -> None:
        class BrokenPageTree:
            is_encrypted = False

            @property
            def pages(self):
                raise ValueError("malformed page tree")

        with patch.object(ingestion, "PdfReader", return_value=BrokenPageTree()):
            self.assert_ingestion_error(self.source("broken.pdf", b"%PDF-1.4\n"), "PDF_PARSE_FAILED")

    def test_pdf_quote_matches_recorded_raw_offsets(self) -> None:
        text = "  Section 3. Cedar  fee\n includes   an override.  \n\n"

        class TextPage:
            def extract_text(self):
                return text

        class Reader:
            is_encrypted = False
            pages = [TextPage()]

        with patch.object(ingestion, "PdfReader", return_value=Reader()):
            result = normalize_file(self.source("raw.pdf", b"%PDF-1.4\n"), DocumentRole.SUPPORTING)
        for item in result.evidence:
            self.assertEqual(item.quote, text[item.text_start : item.text_end])

    def test_upload_limit_is_enforced_before_unbounded_read(self) -> None:
        path = self.source("large.csv", "name\n" + "a" * 100)
        with patch.object(ingestion, "MAX_FILE_BYTES", 8), patch.object(
            Path, "read_bytes", side_effect=AssertionError("unbounded upload read")
        ):
            self.assert_ingestion_error(path, "FILE_TOO_LARGE")

    def test_corrupt_formats_and_encoding_fail_explicitly(self) -> None:
        cases = (
            ("fake.pdf", b"not pdf", "INVALID_PDF"),
            ("truncated.pdf", b"%PDF-1.4\n", "PDF_PARSE_FAILED"),
            ("fake.xlsx", b"not zip", "INVALID_XLSX"),
            ("latin.csv", b"name\ncaf\xe9\n", "CSV_ENCODING"),
        )
        for name, content, code in cases:
            with self.subTest(name=name):
                self.assert_ingestion_error(self.source(name, content), code)

    def test_xlsx_zip_security_and_resource_limits(self) -> None:
        for members, limit_name, limit, code in (
            ({"../outside.xml": b"x"}, None, None, "UNSAFE_XLSX_PACKAGE"),
            ({"a": b"x", "b": b"y"}, "MAX_XLSX_MEMBERS", 1, "XLSX_MEMBER_LIMIT"),
            ({"a": b"12345"}, "MAX_XLSX_UNCOMPRESSED_BYTES", 4, "XLSX_DECOMPRESSED_LIMIT"),
        ):
            with self.subTest(code=code):
                buffer = io.BytesIO()
                with zipfile.ZipFile(buffer, "w") as archive:
                    for name, contents in members.items():
                        archive.writestr(name, contents)
                path = self.source("package.xlsx", buffer.getvalue())
                if limit_name:
                    with patch.object(ingestion, limit_name, limit):
                        self.assert_ingestion_error(path, code)
                else:
                    self.assert_ingestion_error(path, code)

    def test_xlsx_corrupt_central_directory_is_a_typed_error(self) -> None:
        buffer = io.BytesIO()
        with zipfile.ZipFile(buffer, "w") as archive:
            archive.writestr("xl/workbook.xml", "<workbook/>")
        malformed = buffer.getvalue().replace(b"PK\x01\x02", b"BAD!", 1)
        self.assertTrue(zipfile.is_zipfile(io.BytesIO(malformed)))
        self.assert_ingestion_error(self.source("malformed.xlsx", malformed), "INVALID_XLSX")

    def test_csv_row_column_and_cell_limits(self) -> None:
        path = self.source("rows.csv", "name,amount\nCedar,1\nJuniper,2\n")
        for limit_name, limit, code in (
            ("MAX_CSV_ROWS", 1, "CSV_ROW_LIMIT"),
            ("MAX_CSV_COLUMNS", 1, "CSV_COLUMN_LIMIT"),
        ):
            with self.subTest(limit_name=limit_name), patch.object(ingestion, limit_name, limit):
                self.assert_ingestion_error(path, code)
        long_cell = self.source("long.csv", "name\n" + "x" * 20 + "\n")
        with patch.object(ingestion, "MAX_CELL_CHARS", 8):
            result = normalize_file(long_cell, DocumentRole.SUPPORTING)
        self.assertEqual(result.evidence[0].original_value, "x" * 8)
        self.assertEqual(result.document.extraction_status, ExtractionStatus.PARTIAL)

    def test_workbook_dimensions_and_nonempty_cell_limits(self) -> None:
        book, path = self.workbook()
        book.active["A1"] = "party"
        book.active["B2"] = "Cedar"
        book.save(path)
        for limit_name, code in (
            ("MAX_WORKBOOK_ROWS", "WORKSHEET_DIMENSION_LIMIT"),
            ("MAX_WORKBOOK_COLUMNS", "WORKSHEET_DIMENSION_LIMIT"),
            ("MAX_NONEMPTY_CELLS", "WORKBOOK_CELL_LIMIT"),
        ):
            with self.subTest(limit_name=limit_name), patch.object(ingestion, limit_name, 1):
                self.assert_ingestion_error(path, code)

    def test_workbook_grid_budget_prevents_sparse_iteration_and_is_cumulative(self) -> None:
        book, path = self.workbook()
        book.active.title = "first"
        book.active["B2"] = "Cedar"
        book.save(path)
        with self.subTest(case="reject before allocating blank cells"):
            with patch.object(ingestion, "MAX_WORKBOOK_GRID_CELLS", 3, create=True), patch.object(
                Worksheet, "iter_rows", side_effect=AssertionError("must not scan rejected rectangle")
            ) as iterator:
                self.assert_ingestion_error(path, "WORKBOOK_GRID_LIMIT")
                iterator.assert_not_called()

        book.create_sheet("second")["B2"] = "Juniper"
        book.save(path)
        visited_sheets = []
        original_iterator = Worksheet.iter_rows

        def track_iteration(sheet, *args, **kwargs):
            visited_sheets.append(sheet.title)
            return original_iterator(sheet, *args, **kwargs)

        with self.subTest(case="budget sums all sheets"):
            with patch.object(ingestion, "MAX_WORKBOOK_GRID_CELLS", 6, create=True), patch.object(
                Worksheet, "iter_rows", track_iteration
            ):
                self.assert_ingestion_error(path, "WORKBOOK_GRID_LIMIT")
            self.assertEqual(visited_sheets, ["first"])

    def test_pdf_image_only_encrypted_and_page_limits(self) -> None:
        pdf = PdfWriter()
        pdf.add_blank_page(width=612, height=792)
        buffer = io.BytesIO()
        pdf.write(buffer)
        path = self.source("blank.pdf", buffer.getvalue())
        self.assert_ingestion_error(path, "IMAGE_ONLY_OR_EMPTY_PDF")
        with patch.object(ingestion, "MAX_PDF_PAGES", 0):
            self.assert_ingestion_error(path, "PDF_PAGE_LIMIT")
        pdf.encrypt("synthetic-password")
        buffer = io.BytesIO()
        pdf.write(buffer)
        self.assert_ingestion_error(self.source("encrypted.pdf", buffer.getvalue()), "ENCRYPTED_PDF")

    def test_pdf_extraction_failure_and_text_limits(self) -> None:
        class TextPage:
            def extract_text(self):
                return "Cedar rate 1.25 percent."

        class BrokenPage:
            def extract_text(self):
                raise ValueError("broken content stream")

        class Reader:
            is_encrypted = False
            pages = [TextPage(), BrokenPage()]

        path = self.source("partial.pdf", b"%PDF-1.4\n")
        with patch.object(ingestion, "PdfReader", return_value=Reader()):
            result = normalize_file(path, DocumentRole.SUPPORTING)
            self.assertEqual(result.document.extraction_status, ExtractionStatus.PARTIAL)
            self.assertTrue(any("Page 2" in item for item in result.document.warnings))
            with patch.object(ingestion, "MAX_PDF_TOTAL_CHARS", 5):
                self.assert_ingestion_error(path, "PDF_TEXT_LIMIT")
            with patch.object(ingestion, "MAX_PDF_PAGE_CHARS", 5):
                result = normalize_file(path, DocumentRole.SUPPORTING)
                self.assertEqual(result.evidence[0].quote, "Cedar")
                self.assertTrue(any("truncated" in item for item in result.document.warnings))

    def test_three_unseen_source_layouts_preserve_raw_data_and_locations(self) -> None:
        # All three cases differ from the normal NAV/source-pack fixture.
        csv_path = self.source(
            "cedar.csv", '\ufeffAmount GBP;Party;Period\n"(8,765.43)";"  Cedar Research LP  ";2025-Q2\n0;cedar research lp;2025-Q3\n'
        )
        csv_result = normalize_file(csv_path, DocumentRole.SUPPORTING)
        self.assertEqual(csv_result.csv_headers, ["Amount GBP", "Party", "Period"])
        self.assertEqual(csv_result.evidence[1].original_value, "  Cedar Research LP  ")
        self.assertEqual(csv_result.evidence[1].normalized_value, "Cedar Research LP")
        self.assertEqual(csv_result.evidence[3].original_value, "0")

        book, path = self.workbook("cypress.xlsx")
        sheet = book.active
        sheet.title = "Odd Charges Layout"
        sheet.merge_cells("B2:D2")
        sheet["B2"] = "Amounts in £000s"
        sheet["D4"] = "  Cypress Holdings  "
        sheet["B5"] = datetime(2025, 12, 31)
        sheet["D5"] = "2025-12-31"
        sheet["F5"] = -875.5
        sheet["F5"].number_format = '£#,##0.00;[Red](£#,##0.00)'
        sheet["F6"] = "=1/0"
        sheet["F7"] = "#DIV/0!"
        hidden = book.create_sheet("hidden override")
        hidden.sheet_state = "hidden"
        hidden["C9"] = "0.0125"
        book.save(path)
        workbook_result = normalize_file(path, DocumentRole.SUPPORTING)
        by_cell = {(item.sheet, item.cell): item for item in workbook_result.evidence}
        self.assertEqual(by_cell[(sheet.title, "D4")].original_value, "  Cypress Holdings  ")
        self.assertEqual(by_cell[(sheet.title, "F5")].original_value, "-875.5")
        self.assertEqual(by_cell[(sheet.title, "F6")].formula, "=1/0")
        self.assertEqual(by_cell[(sheet.title, "F7")].data_type, "e")
        self.assertIn(("hidden override", "C9"), by_cell)
        self.assertIn("B2:D2", workbook_result.workbook_sheets[0].merged_ranges)
        self.assertTrue(workbook_result.workbook_sheets[1].hidden)

        pdf_buffer = io.BytesIO()
        pdf = canvas.Canvas(pdf_buffer)
        for contents in (
            "Redwood note: candidates GBP 925 and GBP 1,250. Terms absent.",
            "Redwood addendum: rate 1.8 percent; earlier memo says 2.1 percent.",
        ):
            pdf.drawString(50, 700, contents)
            pdf.showPage()
        pdf.save()
        pdf_result = normalize_file(self.source("redwood.pdf", pdf_buffer.getvalue()), DocumentRole.SUPPORTING)
        self.assertEqual({item.page for item in pdf_result.evidence}, {1, 2})
        self.assertTrue(any("2.1 percent" in item.quote for item in pdf_result.evidence))
        self.assertTrue(all(item.document_hash == pdf_result.document.document_hash for item in pdf_result.evidence))


if __name__ == "__main__":
    unittest.main()
