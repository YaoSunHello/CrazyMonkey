"""End-to-end runtime regression tests using files generated and read by ATLAS."""

import csv
from datetime import date
from decimal import Decimal
from pathlib import Path
from tempfile import TemporaryDirectory
import unittest

from openpyxl import load_workbook

from app.atlas import normalize_file
from app.atlas.fixtures import generate_synthetic_pack
from app.atlas.models import NormalizedDocument
from app.runtime.analyst import FixtureAnalyst, ModelAnalyst
from app.runtime.evidence import EvidenceCatalog, build_context
from app.runtime.pipeline import run_case


class _ChangingAnalyst:
    mode = "DEMO_FIXTURE"

    def __init__(self, *, repair=True, nonexistent=False):
        self.calls = []
        self.repair = repair
        self.nonexistent = nonexistent

    def analyse(self, instruction, documents, feedback=None):
        self.calls.append(feedback)
        result = FixtureAnalyst().analyse(instruction, documents, feedback=feedback)
        if len(self.calls) == 1 or not self.repair:
            finding = next(item for item in result.findings if item.investor_id == "LP01")
            if self.nonexistent:
                finding.evidence_ids.append("not_an_atlas_evidence_id")
            else:
                finding.calculation.annual_rate = Decimal("0.015")
        return result


class RuntimePipelineTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.directory = TemporaryDirectory()
        directory = Path(cls.directory.name)
        pack = directory / "full"
        manifest = generate_synthetic_pack(pack)
        cls.documents = [normalize_file(pack / item["filename"], role=item["role"]) for item in manifest["files"]]

        # A smaller PASS case is a perturbation of the canonical fixture sources,
        # not another file parser or fixture generator. Re-normalize changed files
        # through ATLAS so hashes and IDs describe their actual source bytes.
        pass_pack = directory / "pass"
        pass_manifest = generate_synthetic_pack(pass_pack)
        nav = pass_pack / "Administrator_NAV_Q3_2026.xlsx"
        workbook = load_workbook(nav)
        workbook["Investor Fees"].delete_rows(5, 5)
        workbook.save(nav)
        workbook.close()
        register = pass_pack / "investor_input_register.csv"
        with register.open(newline="", encoding="utf-8") as handle:
            rows = list(csv.reader(handle))[:2]
        with register.open("w", newline="", encoding="utf-8") as handle:
            csv.writer(handle).writerows(rows)
        selected = [item for item in pass_manifest["files"] if item["role"] != "SIDE_LETTER" or item["filename"] == "LP01_Side_Letter.pdf"]
        cls.pass_documents = [normalize_file(pass_pack / item["filename"], role=item["role"]) for item in selected]

    @classmethod
    def tearDownClass(cls):
        cls.directory.cleanup()

    def run_pipeline(self, documents=None, analyst=None):
        return run_case("runtime-regression", "Review management fees; prepare PDF and Excel for human review.",
                        self.documents if documents is None else documents, analyst=analyst)

    @staticmethod
    def finding(result, investor_id):
        return next(item for item in result.findings if item.investor_id == investor_id)

    def test_lp03_discrepancy_is_exact_and_requires_human_review(self):
        result = self.run_pipeline()
        finding = self.finding(result, "LP03")
        self.assertEqual(finding.reported, Decimal("50000"))
        self.assertEqual(finding.expected, Decimal("37500.00"))
        self.assertEqual(finding.difference, Decimal("12500.00"))
        self.assertEqual(finding.currency, "GBP")
        self.assertEqual(finding.status, "REVIEW_REQUIRED")
        self.assertEqual(finding.disposition, "NEEDS_HUMAN_REVIEW")
        self.assertEqual(result.repair_count, 1)
        self.assertEqual(result.status, "REVIEW_REQUIRED")
        self.assertTrue(result.output_plan.requires_human_review)

    def test_source_terms_preserve_default_candidate_and_effective_date_metadata(self):
        terms = {item.investor_id: item for item in build_context(EvidenceCatalog(self.documents))}
        self.assertEqual(terms["LP03"].default_annual_rate, Decimal("0.02"))
        self.assertEqual(terms["LP03"].candidate_override_rate, Decimal("0.015"))
        self.assertTrue(terms["LP03"].candidate_override)
        self.assertFalse(terms["LP03"].applicable_default)
        self.assertEqual(terms["LP03"].applicability_state, "APPLIES")
        self.assertEqual(terms["LP03"].effective_from, date(2026, 1, 1))
        self.assertIsNone(terms["LP03"].effective_to)
        self.assertEqual(terms["LP05"].candidate_override_rate, Decimal("0.015"))
        self.assertEqual(terms["LP05"].effective_from, date(2026, 10, 1))
        self.assertTrue(terms["LP05"].candidate_override)
        self.assertTrue(terms["LP05"].applicable_default)
        self.assertEqual(terms["LP05"].applicability_state, "DOES_NOT_APPLY")
        self.assertEqual(terms["LP01"].applicability_state, "APPLIES")
        self.assertTrue(terms["LP01"].applicable_default)
        self.assertFalse(terms["LP01"].candidate_override)
        self.assertEqual(terms["LP06"].applicability_state, "AMBIGUOUS")
        self.assertIsNone(terms["LP06"].applicable_default)

    def test_complete_pass_case_is_verified_without_repair(self):
        result = self.run_pipeline(self.pass_documents)
        self.assertEqual(result.status, "PASS")
        self.assertEqual(result.disposition, "VERIFIED")
        self.assertEqual(result.repair_count, 0)
        self.assertEqual(result.red_team.status, "PASS")
        self.assertFalse(result.output_plan.requires_human_review)
        self.assertEqual(len(result.findings), 1)
        self.assertEqual(self.finding(result, "LP01").difference, Decimal("0.00"))

    def test_missing_expected_side_letter_cannot_pass_after_repair(self):
        result = self.run_pipeline()
        finding = self.finding(result, "LP06")
        self.assertEqual(finding.status, "CANNOT_VERIFY")
        self.assertEqual(finding.disposition, "NEEDS_HUMAN_REVIEW")
        self.assertEqual(finding.reported, Decimal("37500"))
        self.assertIsNone(finding.expected)
        self.assertIsNone(finding.difference)
        self.assertTrue(any(item.insufficient_evidence for item in finding.unresolved_concerns))
        self.assertIn("side letter", finding.explanation.lower())

    def test_single_repair_can_correct_primary_but_preserves_initial_error(self):
        analyst = _ChangingAnalyst()
        result = self.run_pipeline(self.pass_documents, analyst)
        self.assertEqual(len(analyst.calls), 2)
        self.assertIsNone(analyst.calls[0])
        self.assertIn("red_team", analyst.calls[1])
        self.assertIn("verifications", analyst.calls[1])
        self.assertEqual(result.repair_count, 1)
        self.assertEqual(result.initial_red_team.status, "CHALLENGE")
        self.assertEqual(result.initial_analysis.findings[0].calculation.annual_rate, Decimal("0.015"))
        self.assertEqual(result.analysis.findings[0].calculation.annual_rate, Decimal("0.02"))
        self.assertEqual(result.red_team.status, "PASS")
        self.assertEqual(result.status, "PASS")
        self.assertEqual(self.finding(result, "LP01").disposition, "VERIFIED")

    def test_second_wrong_attempt_stops_at_two_calls_and_human_review(self):
        analyst = _ChangingAnalyst(repair=False)
        result = self.run_pipeline(self.pass_documents, analyst)
        self.assertEqual(len(analyst.calls), 2)
        self.assertEqual(result.repair_count, 1)
        self.assertEqual(result.status, "REVIEW_REQUIRED")
        self.assertEqual(result.disposition, "NEEDS_HUMAN_REVIEW")
        self.assertEqual(result.red_team.status, "CHALLENGE")
        self.assertEqual(sum(event.stage == "REPAIRED" for event in result.trace), 1)
        self.assertEqual(self.finding(result, "LP01").expected, Decimal("50000.00"))

    def test_unknown_evidence_is_challenged_and_excluded_from_final_references(self):
        analyst = _ChangingAnalyst(repair=False, nonexistent=True)
        result = self.run_pipeline(self.pass_documents, analyst)
        finding = self.finding(result, "LP01")
        self.assertEqual(finding.status, "REVIEW_REQUIRED")
        self.assertIn("UNKNOWN_EVIDENCE", {item.code for item in finding.unresolved_concerns})
        self.assertNotIn("not_an_atlas_evidence_id", finding.evidence_ids)
        catalog = EvidenceCatalog(self.pass_documents)
        self.assertTrue(all(catalog.validate_ref(ref) for ref in finding.source_refs))
        self.assertEqual(finding.evidence_ids, [ref.evidence_id for ref in finding.source_refs])

    def test_model_cannot_mutate_the_trusted_catalog_via_callback_input(self):
        calls = []
        original = [item.model_dump_json() for item in self.pass_documents]

        def callback(payload):
            calls.append(payload)
            proposals = FixtureAnalyst().analyse(payload["user_instruction"],
                [NormalizedDocument.model_validate(value) for value in payload["normalized_documents"]])
            payload["normalized_documents"][0]["document"]["document_hash"] = "f" * 64
            payload["normalized_documents"][0]["evidence"][0]["quote"] = "Model-mutated source text"
            return proposals.model_dump(mode="json")

        result = self.run_pipeline(self.pass_documents, ModelAnalyst(callback))
        self.assertEqual(result.mode, "MODEL")
        self.assertEqual(len(calls), 1)
        self.assertEqual(result.status, "PASS")
        self.assertEqual([item.model_dump_json() for item in self.pass_documents], original)
        catalog = EvidenceCatalog(self.pass_documents)
        self.assertTrue(all(catalog.validate_ref(ref) for finding in result.findings for ref in finding.source_refs))

    def test_changed_source_hash_is_rejected_at_catalog_entry(self):
        documents = [item.model_copy(deep=True) for item in self.pass_documents]
        documents[0].document.document_hash = "f" * 64
        analyst = _ChangingAnalyst()
        with self.assertRaises(ValueError):
            self.run_pipeline(documents, analyst)
        self.assertEqual(analyst.calls, [])

    def test_operational_trace_has_exact_zero_repair_order(self):
        result = self.run_pipeline(self.pass_documents)
        self.assertEqual([event.stage for event in result.trace], [
            "INGESTED", "ANALYSED", "RED_TEAMED", "VERIFIED", "FINAL_VERIFIED", "OUTPUT_PLANNED"])
        self.assertTrue(all(event.timestamp.tzinfo is not None for event in result.trace))
        self.assertEqual([event.timestamp for event in result.trace], sorted(event.timestamp for event in result.trace))
        self.assertTrue(all(event.explanation for event in result.trace))

    def test_operational_trace_has_exact_one_repair_order(self):
        result = self.run_pipeline(self.pass_documents, _ChangingAnalyst())
        self.assertEqual([event.stage for event in result.trace], [
            "INGESTED", "ANALYSED", "RED_TEAMED", "VERIFIED", "REPAIRED", "RED_TEAMED", "FINAL_VERIFIED", "OUTPUT_PLANNED"])

    def test_empty_and_unknown_source_scope_never_become_pass(self):
        unknown_documents = [item.model_copy(deep=True) for item in self.pass_documents]
        for document in unknown_documents:
            document.document.role = "SUPPORTING"
        for documents in ([], unknown_documents):
            with self.subTest(documents=len(documents)):
                result = self.run_pipeline(documents)
                self.assertEqual(result.status, "CANNOT_VERIFY")
                self.assertEqual(result.disposition, "NEEDS_HUMAN_REVIEW")
                self.assertTrue(result.findings)
                self.assertTrue(all(item.status == "CANNOT_VERIFY" for item in result.findings))
                self.assertEqual(result.repair_count, 1)

    def test_model_exception_fails_closed_without_exposing_provider_error(self):
        calls = []

        def unavailable(payload):
            calls.append(payload)
            raise RuntimeError("SENSITIVE_PROVIDER_RESPONSE_NOT_FOR_PUBLIC_OUTPUT")

        result = self.run_pipeline(self.pass_documents, ModelAnalyst(unavailable))
        self.assertEqual(len(calls), 2)
        self.assertNotEqual(result.status, "PASS")
        self.assertEqual(result.disposition, "NEEDS_HUMAN_REVIEW")
        self.assertEqual(result.mode, "MODEL")
        self.assertNotIn("SENSITIVE_PROVIDER_RESPONSE_NOT_FOR_PUBLIC_OUTPUT", result.model_dump_json())
        self.assertTrue(any("failed" in item.lower() for item in result.limitations))

    def test_unvalidated_model_response_cannot_become_a_pass(self):
        def invalid_response(payload):
            return {"findings": [], "private_reasoning": "This field must be rejected"}

        result = self.run_pipeline(self.pass_documents, ModelAnalyst(invalid_response))
        self.assertNotEqual(result.status, "PASS")
        self.assertNotIn("private_reasoning", result.model_dump_json())
        self.assertEqual(result.repair_count, 1)

    def test_demo_mode_is_explicit_and_output_plan_does_not_send_email(self):
        result = self.run_pipeline(self.pass_documents)
        self.assertEqual(result.mode, "DEMO_FIXTURE")
        self.assertTrue(any("no language model was invoked" in item for item in result.limitations))
        self.assertEqual(result.output_plan.delivery, "RETURN_TO_CALLER")
        self.assertEqual({item.format for item in result.output_plan.recommendations}, {"PDF", "XLSX", "JSON"})
        self.assertIn("explicit confirmation", result.output_plan.explanation)

    def test_unsupported_workbook_notes_and_adjustment_columns_fail_closed(self):
        for variant in ("hidden_note", "adjustment_column"):
            with self.subTest(variant=variant), TemporaryDirectory() as name:
                directory = Path(name)
                manifest = generate_synthetic_pack(directory)
                path = directory / "Administrator_NAV_Q3_2026.xlsx"
                workbook = load_workbook(path)
                if variant == "hidden_note":
                    notes = workbook.create_sheet("Fee qualification")
                    notes["A1"] = "LP01 management fee is waived for this quarter."
                    notes.sheet_state = "hidden"
                else:
                    workbook["Investor Fees"]["I3"] = "Fee Adjustment"
                    workbook["Investor Fees"]["I4"] = 100
                workbook.save(path)
                workbook.close()
                documents = [normalize_file(directory / item["filename"], role=item["role"]) for item in manifest["files"]]
                result = self.run_pipeline(documents)
                self.assertEqual(self.finding(result, "LP01").status, "CANNOT_VERIFY")
                self.assertTrue(all(item.status != "PASS" for item in result.findings))

    def test_unsupported_register_adjustment_column_is_not_ignored(self):
        with TemporaryDirectory() as name:
            directory = Path(name)
            manifest = generate_synthetic_pack(directory)
            path = directory / "investor_input_register.csv"
            with path.open(newline="", encoding="utf-8") as handle:
                rows = list(csv.reader(handle))
            rows[0].append("fee_base_adjustment")
            for row in rows[1:]:
                row.append("100")
            with path.open("w", newline="", encoding="utf-8") as handle:
                csv.writer(handle).writerows(rows)
            documents = [normalize_file(directory / item["filename"], role=item["role"]) for item in manifest["files"]]
            result = self.run_pipeline(documents)
            self.assertEqual(self.finding(result, "LP01").status, "CANNOT_VERIFY")
            self.assertTrue(all(item.status != "PASS" for item in result.findings))


if __name__ == "__main__":
    unittest.main()
