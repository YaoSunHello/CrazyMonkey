"""Real ATLAS/Decimal/openpyxl integration, with explicitly mocked model transport."""
import asyncio
import copy
import tempfile
import threading
import time
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

from app.atlas.fixtures import generate_synthetic_pack
from app.fast_audit import _deduplicate, _ingest, from_plan, run_audit
from app.runtime.fast_dsl import execute_check
from app.runtime.model import RuntimeModelError
from app.runtime.planner import offline_plan


class ParallelModel:
    name = "TEST_ONLY_GEMINI_MOCK"

    def __init__(self, plans, *, invalid=False, failure=False, conflict=False):
        self.plans = plans
        self.invalid, self.failure, self.conflict = invalid, failure, conflict
        self.barrier = threading.Barrier(2)
        self.calls = []
        self.lock = threading.Lock()

    def complete_json(self, system, payload, *, stage):
        if stage in {"contract_discovery", "relationship_discovery"}:
            self.barrier.wait(timeout=5)
        if self.failure:
            raise RuntimeModelError("mock transport unavailable")
        with self.lock:
            self.calls.append({"stage": stage, "status": "success", "provider": "gemini", "response_id": "test-only"})
        if stage == "red_team":
            return {"status": "PASS", "reasons": [], "evidence_ids": payload["check"]["context_evidence_ids"]}
        if self.invalid:
            return {"checks": [{"arbitrary_python": "prohibited"}]}
        if stage == "relationship_discovery":
            if self.conflict:
                check = copy.deepcopy(next(plan for plan in self.plans if plan["entity_id"] == "LP03"))
                check["operation"] = "ADD"
                return {"checks": [check]}
            return {"checks": []}
        return {"checks": copy.deepcopy(self.plans)}


class FastAuditTests(unittest.TestCase):
    def setUp(self):
        self.temp = tempfile.TemporaryDirectory()
        self.root = Path(self.temp.name)
        self.input = self.root / "pack"
        generate_synthetic_pack(self.input)
        self.store, _, _ = _ingest(self.input)
        self.plans = [from_plan(plan).model_dump(mode="json") for plan in offline_plan(self.store).checks]

    def tearDown(self):
        self.temp.cleanup()

    def run_case(self, **kwargs):
        return asyncio.run(run_audit(self.input, "Find and repair material discrepancies.",
                                     output_dir=self.root / "out", **kwargs))[0]

    def test_parallel_discovery_verification_selective_red_team_and_patch(self):
        model = ParallelModel(self.plans)
        workbook = self.input / "Administrator_NAV_Q3_2026.xlsx"
        before = workbook.read_bytes()
        with patch("app.fast_audit.execute_check", side_effect=lambda *a, **kw: (time.sleep(.02), execute_check(*a, **kw))[1]), \
                patch("app.fast_audit.normalize_file", wraps=__import__("app.fast_audit", fromlist=["normalize_file"]).normalize_file) as normalize:
            result = self.run_case(model=model)
        self.assertEqual(normalize.call_count, 8)
        fee = next(f for f in result["findings"] if f["entity_id"] == "LP03" and f["check"]["check_type"] == "annual_charge")
        self.assertEqual(fee["status"], "DISCREPANCY")
        self.assertEqual(fee["calculation"]["expected"], "37500.00")
        self.assertEqual(fee["calculation"]["difference"], "12500.00")
        self.assertGreaterEqual(result["peak_concurrency"]["investigation"], 2)
        self.assertGreaterEqual(result["peak_concurrency"]["verification"], 2)
        self.assertGreaterEqual(result["peak_concurrency"]["red_team"], 2)
        red_calls = [call for call in model.calls if call["stage"] == "red_team"]
        self.assertEqual(len(red_calls), 3)  # LP03 material; LP04/LP06 ambiguous. Clean records skipped.
        self.assertEqual(workbook.read_bytes(), before)
        patched = load_workbook(result["patches"][0]["output_file"])
        self.assertEqual(patched["Investor Fees"]["F6"].value, 37500)
        self.assertIn("Audit Trail", patched.sheetnames)
        patched.close()
        for entity in ("LP04", "LP06"):
            self.assertEqual(next(f for f in result["findings"] if f["entity_id"] == entity)["status"], "CANNOT_VERIFY")

    def test_transport_failure_never_falls_back_or_patches(self):
        with patch("app.fast_audit.offline_plan", side_effect=AssertionError("fallback forbidden")):
            with self.assertRaisesRegex(RuntimeModelError, "mock transport"):
                self.run_case(model=ParallelModel(self.plans, failure=True))
        self.assertFalse((self.root / "out").exists())

    def test_invalid_structured_batch_gets_one_repair_then_no_patch(self):
        model = ParallelModel(self.plans, invalid=True)
        result = self.run_case(model=model)
        self.assertTrue(result["repair_attempted"])
        self.assertEqual([call["stage"] for call in model.calls].count("repair"), 1)
        self.assertTrue(result["cannot_verify"])
        self.assertEqual(result["patches"], [])

    def test_conflicting_target_relationships_do_not_patch(self):
        result = self.run_case(model=ParallelModel(self.plans, conflict=True))
        lp03 = [f for f in result["findings"] if f["entity_id"] == "LP03" and f["check"]["compare_to"]]
        self.assertTrue(lp03)
        self.assertTrue(all(f["status"] == "CANNOT_VERIFY" for f in lp03))
        self.assertEqual(result["patches"], [])

    def test_model_cannot_override_executor_disagreement(self):
        def disagree(check, store, tolerance):
            result = execute_check(check, store, tolerance)
            if check.entity_id == "LP03" and check.compare_to:
                result["expected"] = "1.00"
            return result
        with patch("app.fast_audit.execute_check", side_effect=disagree):
            result = self.run_case(model=ParallelModel(self.plans))
        fee = next(f for f in result["findings"] if f["entity_id"] == "LP03" and f["check"]["compare_to"])
        self.assertEqual(fee["status"], "CANNOT_VERIFY")
        self.assertIsNone(fee["model_review"])
        self.assertEqual(result["patches"], [])

    def test_no_apply_flag_leaves_only_reviewable_proposals(self):
        result = self.run_case(mode="SYNTHETIC_DEMO", apply_verified_fixes=False)
        self.assertEqual(result["patches"], [])
        self.assertTrue(any(f["patch_proposal"] for f in result["findings"]))
        self.assertEqual(result["gemini_call_count"], 0)
        self.assertFalse((self.root / "out").exists())

    def test_deduplication_merges_context_but_preserves_stricter_template(self):
        first = from_plan(offline_plan(self.store).checks[0])
        second = first.model_copy(update={"check_id": "different", "check_type": "model_proposed", "inputs": list(reversed(first.inputs))})
        checks, conflicts = _deduplicate([second, first])
        self.assertEqual(len(checks), 1)
        self.assertEqual(checks[0].check_type, "annual_charge")
        self.assertEqual(conflicts, set())


if __name__ == "__main__":
    unittest.main()
