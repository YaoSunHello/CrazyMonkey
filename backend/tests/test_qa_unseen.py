"""Different source layouts exercise the real normalizer, not demo answer JSON.

These tests establish extraction/provenance only. There is no implemented fee
review pipeline to assert investor selection, contractual applicability or fees.
All entities and documents generated here are fictional.
"""

from __future__ import annotations

import csv
import tempfile
import unittest
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from pypdf import PdfReader
from reportlab.pdfgen import canvas

from app.atlas.ingestion import normalize_file
from app.atlas.models import DocumentRole, NormalizedDocument


def _pdf(path: Path, pages: list[list[str]]) -> None:
    document = canvas.Canvas(str(path), invariant=1)
    document.setAuthor("CrazyMonkey QA synthetic fixtures")
    for lines in pages:
        for index, line in enumerate(lines):
            document.drawString(36, 780 - index * 22, line)
        document.showPage()
    document.save()


def generate_unseen_cases(output_dir: Path) -> list[dict]:
    """Create three packs with changed terminology, layout, identity and order."""
    cases = []
    for case_id in ("sterling_thousands", "dollar_vertical", "euro_reordered"):
        directory = output_dir / case_id
        directory.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        register = directory / "clients.csv"
        statement = directory / "statement.xlsx"
        agreement = directory / "terms.pdf"
        if case_id == "sterling_thousands":
            workbook.active.title = "Empty cover"
            sheet = workbook.create_sheet("Capital Accounts")
            sheet.merge_cells("A1:F1")
            sheet["A1"] = "FICTIONAL Alder Test Fund / all money in GBP thousands"
            sheet.append([])
            for cell, value in {
                "A3": "Client reference", "C3": "Assessed charge (GBP 000s)",
                "D3": "Subscribed capital (GBP 000s)", "E3": "CCY", "F3": "Close date",
                "A4": "AX-17", "C4": 31.5, "D4": 8400, "E4": "GBP",
                "F4": datetime(2027, 6, 30),
            }.items():
                sheet[cell] = value
            sheet["C4"].number_format = '£0.000'
            notes = workbook.create_sheet("Administrator notes")
            notes.sheet_state = "hidden"
            notes["A1"] = "Amounts above are in thousands, not whole pounds."
            rows = [["Client reference", "Display name", "CCY"],
                    ["AX-17", "Alder QA Trust", "GBP"],
                    ["AX-18", "Alder QA Trust", "GBP"]]
            pages = [["FICTIONAL Alder Test Fund - Q2 2027", "Annual advisory charge: 1.50%.",
                      "AX-17 has an annual override of 1.25% effective 1 January 2027."],
                     ["FICTIONAL Alder Test Fund - Q2 2027", "Quarterly multiplier: 0.25.",
                      "Capital: GBP 8,400,000. Assessed charge: GBP 31,500."]]
            role = DocumentRole.NAV_WORKBOOK
            order = [register, statement, agreement]
            anchors = {("Capital Accounts", "A4"): "AX-17",
                       ("Capital Accounts", "C4"): "31.5",
                       ("Capital Accounts", "D4"): "8400"}
        elif case_id == "dollar_vertical":
            sheet = workbook.active
            sheet.title = "Vertical Statement"
            for row in [["FICTIONAL Birch QA Fund", "Account summary"],
                        ["Account", "BX-92"], ["Account capital", "$2,750,000.00"],
                        ["Charge booked", "(13,750.00)"], ["Currency", "USD"],
                        ["Closing date", "2028-03-31"], ["Annual percentage", "2.00%"],
                        ["Spreadsheet check", "=1/0"], ["Administrator error", "#DIV/0!"],
                        ["Zero adjustment", 0]]:
                sheet.append(row)
            notes = workbook.create_sheet("Prior terms")
            notes.sheet_state = "veryHidden"
            notes["A1"] = "Future BX-92 term: 1.00% from 2028-04-01."
            rows = [["Currency", "Account holder", "Account"],
                    ["USD", "Birch, QA Endowment", "BX-92"]]
            pages = [["FICTIONAL Birch QA Fund. Account BX-92.",
                      "Management levy: 2.00% per annum. Review quarter: Q1 2028.",
                      "A 1.00% concession starts 1 April 2028, after this review quarter."]]
            role = DocumentRole.NAV_WORKBOOK
            order = [agreement, register, statement]
            anchors = {("Vertical Statement", "B2"): "BX-92",
                       ("Vertical Statement", "B3"): "$2,750,000.00",
                       ("Vertical Statement", "B4"): "(13,750.00)",
                       ("Vertical Statement", "B10"): "0"}
        else:
            sheet = workbook.active
            sheet.title = "OtherReturn"
            for row in [["FICTIONAL Cedar QA Fund", None, None, None, None],
                        ["Currency", "Expense charged", "Member", "Valuation", "Period end"],
                        ["EUR", "€9,187.50", "cx-504", "3,500,000", "30/09/2029"],
                        ["EUR", "0.00", "CX-505", 0, datetime(2029, 9, 30)]]:
                sheet.append(row)
            rows = [["Member", "Display name", "Agreement expected"],
                    ["cx-504", "Cedar QA Trust", "missing_concession.pdf"],
                    ["CX-505", "CEDAR QA TRUST", "none"]]
            pages = [["FICTIONAL Cedar QA Fund. Members CX-504 and CX-505.",
                      "Headline annual rate: 1.40%. Another clause states 1.05%.",
                      "No governing priority or concession agreement has been supplied."]]
            role = DocumentRole.NAV_WORKBOOK
            order = [statement, agreement, register]
            anchors = {("OtherReturn", "B3"): "€9,187.50",
                       ("OtherReturn", "C3"): "cx-504",
                       ("OtherReturn", "D3"): "3,500,000",
                       ("OtherReturn", "D4"): "0"}
        workbook.save(statement)
        workbook.close()
        with register.open("w", encoding="utf-8-sig", newline="") as handle:
            csv.writer(handle, delimiter=";" if case_id == "dollar_vertical" else ",").writerows(rows)
        _pdf(agreement, pages)
        cases.append({"case_id": case_id, "paths": order, "anchors": anchors,
                      "workbook_role": role, "pdf_pages": len(pages)})
    return cases


class UnseenInputTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp = tempfile.TemporaryDirectory()
        self.addCleanup(self.temp.cleanup)
        self.cases = generate_unseen_cases(Path(self.temp.name))

    def _check_case(self, case: dict) -> None:
        normalized = []
        for path in case["paths"]:
            # Unknown filenames must remain visibly unconfirmed by default.
            extracted = normalize_file(path)
            self.assertEqual(extracted.document.extraction_status, "NEEDS_CONFIRMATION")
            self.assertTrue(extracted.evidence)
            self.assertEqual(extracted.document.filename, path.name)
            self.assertEqual(NormalizedDocument.model_validate_json(
                extracted.model_dump_json()).model_dump(), extracted.model_dump())
            self.assertTrue(all(ref.document_id == extracted.document.document_id
                                and ref.document_hash == extracted.document.document_hash
                                for ref in extracted.evidence))
            normalized.append(extracted)
        book = next(item for item in normalized if item.document.filename.endswith(".xlsx"))
        cells = {(ref.sheet, ref.cell): ref for ref in book.evidence}
        for locator, value in case["anchors"].items():
            self.assertEqual(cells[locator].original_value, value)
        source_book = load_workbook(next(p for p in case["paths"] if p.suffix == ".xlsx"))
        try:
            self.assertEqual(source_book.sheetnames, [sheet.name for sheet in book.workbook_sheets])
        finally:
            source_book.close()
        pdf = next(item for item in normalized if item.document.filename.endswith(".pdf"))
        reader = PdfReader(next(p for p in case["paths"] if p.suffix == ".pdf"))
        self.assertEqual(pdf.layout["page_count"], case["pdf_pages"])
        self.assertEqual(len(reader.pages), case["pdf_pages"])
        self.assertTrue(all(ref.page is not None and ref.quote for ref in pdf.evidence))
        # Order changes must not affect per-file normalized evidence identity.
        reversed_by_name = {path.name: normalize_file(path) for path in reversed(case["paths"])}
        for original in normalized:
            self.assertEqual(original, reversed_by_name[original.document.filename])

    def test_sterling_thousands_merged_headers_duplicate_names(self) -> None:
        self._check_case(self.cases[0])

    def test_dollar_vertical_future_term_formulas_and_credit(self) -> None:
        self._check_case(self.cases[1])

    def test_euro_reordered_conflicting_terms_missing_agreement(self) -> None:
        self._check_case(self.cases[2])


if __name__ == "__main__":
    unittest.main()
