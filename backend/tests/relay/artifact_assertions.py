from __future__ import annotations

from email import policy
from email.message import EmailMessage
from email.parser import BytesParser
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence
from zipfile import ZipFile

from openpyxl.worksheet.worksheet import Worksheet
from pypdf import PdfReader


def compact_text(value: Any) -> str:
    """Return display text with PDF/Excel whitespace normalised."""

    if value is None:
        return ""
    return " ".join(str(value).replace("\u00a0", " ").split())


def casefolded_text(value: Any) -> str:
    return compact_text(value).casefold()


def pdf_text(path: Path) -> str:
    reader = PdfReader(path)
    assert reader.pages, "generated PDF has no pages"
    return "\n".join(page.extract_text() or "" for page in reader.pages)


def all_cell_values(sheet: Worksheet) -> Iterable[Any]:
    for row in sheet.iter_rows():
        for cell in row:
            yield cell.value


def find_labeled_value(sheet: Worksheet, label: str) -> Any:
    """Find a value immediately to the right of a case-insensitive label."""

    wanted = casefolded_text(label).removesuffix(":")
    for row in sheet.iter_rows():
        for index, cell in enumerate(row[:-1]):
            candidate = casefolded_text(cell.value).removesuffix(":")
            if candidate == wanted:
                return row[index + 1].value
    raise AssertionError(f"label {label!r} not found on sheet {sheet.title!r}")


def table_records(
    sheet: Worksheet,
    required_headers: Sequence[str],
) -> list[dict[str, Any]]:
    """Locate a table by headers and return its non-empty rows as dictionaries."""

    required = {casefolded_text(header) for header in required_headers}
    for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        normalised = [casefolded_text(value) for value in row]
        present = {value for value in normalised if value}
        if not required.issubset(present):
            continue

        header_by_column = {
            column_index: compact_text(value)
            for column_index, value in enumerate(row, start=1)
            if compact_text(value)
        }
        records: list[dict[str, Any]] = []
        for values in sheet.iter_rows(
            min_row=row_index + 1,
            max_col=max(header_by_column),
            values_only=True,
        ):
            record = {
                header: values[column_index - 1]
                for column_index, header in header_by_column.items()
            }
            if any(value not in (None, "") for value in record.values()):
                records.append(record)
        return records

    raise AssertionError(
        f"headers {list(required_headers)!r} not found on sheet {sheet.title!r}"
    )


def record_matching(
    records: Iterable[Mapping[str, Any]],
    key: str,
    expected: Any,
) -> Mapping[str, Any]:
    wanted = casefolded_text(expected)
    for record in records:
        actual_key = next(
            (candidate for candidate in record if casefolded_text(candidate) == casefolded_text(key)),
            None,
        )
        if actual_key is not None and casefolded_text(record[actual_key]) == wanted:
            return record
    raise AssertionError(f"no table record has {key!r}={expected!r}")


def record_value(record: Mapping[str, Any], header: str) -> Any:
    wanted = casefolded_text(header)
    for candidate, value in record.items():
        if casefolded_text(candidate) == wanted:
            return value
    raise AssertionError(f"header {header!r} absent from {list(record)!r}")


def parse_eml(path: Path) -> EmailMessage:
    message = BytesParser(policy=policy.default).parsebytes(path.read_bytes())
    assert isinstance(message, EmailMessage)
    return message


def email_body(message: EmailMessage) -> str:
    body = message.get_body(preferencelist=("plain",))
    if body is None:
        payload = message.get_payload(decode=True)
        return payload.decode(message.get_content_charset() or "utf-8") if payload else ""
    return body.get_content()


def attachment_filenames(message: EmailMessage) -> list[str]:
    return [part.get_filename() or "" for part in message.iter_attachments()]


def assert_xlsx_has_no_active_or_external_content(path: Path) -> None:
    """Reject macros and OOXML external-link relationships in a generated workbook."""

    with ZipFile(path) as archive:
        names = [name.casefold() for name in archive.namelist()]
        assert not any(name.endswith("vbaproject.bin") for name in names)
        assert not any("externallinks/" in name for name in names)

        for name in archive.namelist():
            if not name.casefold().endswith(".rels"):
                continue
            relationships = archive.read(name).lower()
            assert b'targetmode="external"' not in relationships


def assert_close_money(actual: Any, expected: float) -> None:
    assert isinstance(actual, (int, float)), f"expected numeric money cell, got {actual!r}"
    assert float(actual) == expected
