from __future__ import annotations

import copy
import json
from email import policy
from email.parser import BytesParser
from pathlib import Path

import pytest
from openpyxl import load_workbook
from pypdf import PdfReader

from app.atlas.models import ReviewSnapshot
from app.relay.contracts import SnapshotContractError, adapt_review_snapshot
from app.relay.export_service import ExportService
from app.relay.models import FindingStatus, ReviewMode


def _atlas_payload() -> dict[str, object]:
    document_hash = "a" * 64
    source_ref = {
        "evidence_id": "ev-nav-lp03-fee",
        "document_id": "doc-nav",
        "document_hash": document_hash,
        "kind": "WORKBOOK_CELL",
        "sheet": "Investor Fees",
        "cell": "F6",
        "original_value": "50000.00",
        "normalized_value": "50000.00",
        "data_type": "number",
        "number_format": "GBP #,##0.00",
    }
    return {
        "contract_version": 1,
        "run_id": "atlas-review-q3-2026",
        "version": 3,
        "mode": "SYNTHETIC_DEMO",
        "fund_name": "Example Growth Fund III",
        "reporting_period": "Q3 2026",
        "created_at": "2026-09-05T10:00:00Z",
        "frozen_at": "2026-09-05T10:05:00Z",
        "source_documents": [
            {
                "document_id": "doc-nav",
                "filename": "Administrator_NAV_Q3_2026.xlsx",
                "document_hash": document_hash,
                "role": "NAV_WORKBOOK",
                "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "size_bytes": 4096,
                "extraction_status": "COMPLETE",
                "warnings": [],
                "original_storage_key": "runs/atlas-review-q3-2026/doc-nav",
            }
        ],
        "rules": [
            {
                "rule_id": "rule-lp03-fee",
                "rule_version": 1,
                "investor_id": "LP03",
                "identity_evidence_ids": ["ev-nav-lp03-fee"],
                "term_type": "MANAGEMENT_FEE",
                "default_annual_rate": "0.02",
                "candidate_override_rate": "0.015",
                "applicable_annual_rate": "0.015",
                "fee_basis": "Committed capital",
                "fee_base": "10000000.00",
                "currency": "GBP",
                "period_start": "2026-07-01",
                "period_end": "2026-09-30",
                "period_factor": "0.25",
                "effective_from": "2026-01-01",
                "applicable_default": False,
                "candidate_override": True,
                "applicability_state": "APPLIES",
                "applicability_rationale": "The investor-specific rate is effective for Q3 2026.",
                "input_evidence": {
                    "identity": ["ev-nav-lp03-fee"],
                    "fee_base": ["ev-nav-lp03-fee"],
                },
                "extraction_state": "SUPPORTED",
                "unresolved_questions": [],
            }
        ],
        "calculations": [
            {
                "calculation_id": "calc-lp03-fee",
                "rule_id": "rule-lp03-fee",
                "rule_version": 1,
                "investor_id": "LP03",
                "formula_code": "ANNUAL_RATE_X_PERIOD_FACTOR_X_FEE_BASE",
                "formula_description": "GBP 10,000,000 x 1.5% x 0.25 = GBP 37,500",
                "fee_base": "10000000.00",
                "annual_rate": "0.015",
                "period_factor": "0.25",
                "currency": "GBP",
                "rounding": "ROUND_HALF_UP_0.01",
                "tolerance": "0.01",
                "expected_amount": "37500.00",
                "reported_amount": "50000.00",
                "difference": "12500.00",
                "input_evidence": {"reported_fee": ["ev-nav-lp03-fee"]},
            }
        ],
        "challenger_concerns": [
            {
                "concern_id": "concern-lp03-rate",
                "investor_id": "LP03",
                "rule_id": "rule-lp03-fee",
                "severity": "WARNING",
                "state": "RESOLVED",
                "suspected_problem": "The administrator may have used the LPA default rate.",
                "evidence_ids": ["ev-nav-lp03-fee"],
                "required_resolution": "Confirm the effective investor rate.",
            }
        ],
        "verifier_results": [
            {
                "verifier_result_id": "verify-lp03-fee",
                "investor_id": "LP03",
                "rule_id": "rule-lp03-fee",
                "calculation_id": "calc-lp03-fee",
                "status": "DISCREPANCY",
                "checks": [
                    {
                        "code": "ARITHMETIC",
                        "passed": True,
                        "explanation": "The expected amount recomputes to GBP 37,500.",
                    }
                ],
                "blocking_concern_ids": [],
                "explanation": "Reported and expected fees differ by GBP 12,500.",
            }
        ],
        "findings": [
            {
                "finding_id": "finding-lp03-fee",
                "investor_id": "LP03",
                "check_type": "MANAGEMENT_FEE",
                "reported_value": "50000.00",
                "expected_value": "37500.00",
                "difference": "12500.00",
                "currency": "GBP",
                "status": "DISCREPANCY",
                "human_review_state": "UNREVIEWED",
                "calculation_id": "calc-lp03-fee",
                "evidence_ids": ["ev-nav-lp03-fee"],
                "source_refs": [source_ref],
                "challenger_concern_ids": ["concern-lp03-rate"],
                "verifier_result_id": "verify-lp03-fee",
                "explanation": "The reported fee used a higher rate than the supported rule.",
                "actionable_next_step": "Review the supported override and administrator entry.",
                "unresolved_questions": [],
            }
        ],
        "audit_trail": [],
        "unresolved_items": [],
        "limitations": ["Management-fee checks only.", "Synthetic fixture; no regulated sign-off."],
    }


def test_atlas_review_snapshot_v1_maps_without_recomputing_values() -> None:
    atlas = ReviewSnapshot.model_validate(_atlas_payload())
    snapshot = adapt_review_snapshot(atlas.model_dump(mode="json"))

    assert snapshot.run_id == "atlas-review-q3-2026"
    assert snapshot.version == 3
    assert snapshot.mode == ReviewMode.SYNTHETIC_DEMO
    assert snapshot.timestamp.isoformat() == "2026-09-05T10:05:00+00:00"
    assert snapshot.source == "ATLAS_REVIEW_SNAPSHOT_V1"
    assert snapshot.summary_counts()["discrepancies"] == 1
    assert snapshot.limitations == (
        "Management-fee checks only.",
        "Synthetic fixture; no regulated sign-off.",
    )

    finding = snapshot.findings[0]
    assert finding.computational_status == FindingStatus.DISCREPANCY
    assert finding.administrator_value == 50_000
    assert finding.expected_value == 37_500
    assert finding.difference == 12_500
    assert finding.calculation_id == "calc-lp03-fee"

    assert snapshot.source_documents[0].sha256 == "a" * 64
    assert snapshot.evidence_references[0].locator == "Investor Fees!F6"
    assert snapshot.investor_terms[0].applicable_rate_fraction == 0.015
    assert snapshot.calculations[0].expected_value == 37_500
    assert snapshot.challenger_concerns[0].finding_id == finding.finding_id
    assert snapshot.verifier_results[0].finding_id == finding.finding_id


def test_atlas_live_modes_are_explicitly_projected_as_live() -> None:
    payload = _atlas_payload()
    payload["mode"] = "LIVE_OFFLINE"

    snapshot = adapt_review_snapshot(payload)

    assert snapshot.mode == ReviewMode.LIVE
    assert snapshot.source_notice == (
        "Validated Atlas ReviewSnapshot v1; upstream mode: LIVE_OFFLINE."
    )


def test_atlas_source_reference_hash_mismatch_is_rejected() -> None:
    payload = copy.deepcopy(_atlas_payload())
    payload["findings"][0]["source_refs"][0]["document_hash"] = "b" * 64  # type: ignore[index]

    with pytest.raises(SnapshotContractError, match="hash"):
        adapt_review_snapshot(payload)


def test_atlas_snapshot_drives_one_consistent_pdf_xlsx_json_and_eml_bundle(
    tmp_path: Path,
    export_schema_path: Path,
) -> None:
    atlas = ReviewSnapshot.model_validate(_atlas_payload())
    service = ExportService(tmp_path / "relay-output", export_schema_path)
    frozen = service.snapshot_store.freeze(atlas.model_dump(mode="json"))
    bundle = service.generate_all(frozen)

    assert (bundle.run_id, bundle.version, bundle.snapshot_sha256) == frozen.identity
    assert {item.artifact_type for item in bundle.artifacts} == {
        "pdf",
        "xlsx",
        "json",
        "eml",
    }

    public_json = json.loads(bundle.artifact_path("json").read_text(encoding="utf-8"))
    assert public_json["export_metadata"]["run_id"] == atlas.run_id
    assert public_json["export_metadata"]["version"] == atlas.version
    assert public_json["export_metadata"]["snapshot_sha256"] == frozen.snapshot_sha256
    assert public_json["findings"][0]["expected_value"] == 37_500
    assert public_json["source_documents"][0]["sha256"] == "a" * 64
    assert public_json["limitations"] == atlas.limitations

    pdf_text = "\n".join(
        page.extract_text() or ""
        for page in PdfReader(str(bundle.artifact_path("pdf"))).pages
    )
    assert atlas.run_id in pdf_text
    assert "£37,500.00" in pdf_text
    assert "Upstream limitation: Management-fee checks only." in pdf_text

    workbook = load_workbook(bundle.artifact_path("xlsx"), data_only=False)
    assert workbook.sheetnames == [
        "Summary",
        "Findings",
        "Investor Terms",
        "Calculations",
        "Sources",
        "Audit Trail",
    ]
    assert any(
        cell.value == atlas.run_id
        for row in workbook["Summary"].iter_rows()
        for cell in row
    )

    draft = BytesParser(policy=policy.default).parsebytes(
        bundle.artifact_path("eml").read_bytes()
    )
    assert draft["To"] is None
    assert draft["X-YLookup-Run-ID"] == atlas.run_id
    assert draft["X-YLookup-Review-Version"] == str(atlas.version)
    assert str(draft["X-YLookup-Snapshot-SHA256"]).strip() == frozen.snapshot_sha256
    assert sorted(part.get_filename() for part in draft.iter_attachments()) == sorted(
        item.filename
        for item in bundle.artifacts
        if item.artifact_type in {"pdf", "xlsx", "json"}
    )
