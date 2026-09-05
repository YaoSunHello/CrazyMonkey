from __future__ import annotations

import copy
import re
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

from app.relay.json_export import build_review_export
from app.relay.snapshot_store import FileSnapshotStore, FrozenSnapshot
from app.relay.xlsx_export import SHEET_NAMES, write_xlsx_report
from .artifact_assertions import (
    assert_close_money,
    assert_xlsx_has_no_active_or_external_content,
    record_matching,
    record_value,
    table_records,
)


GENERATED_AT = datetime(2026, 9, 5, 12, 0, tzinfo=timezone.utc)


def test_xlsx_opens_with_required_sheets_values_formulas_and_audit(
    tmp_path: Path,
    frozen_snapshot: FrozenSnapshot,
) -> None:
    path = tmp_path / "review.xlsx"
    public_export = build_review_export(frozen_snapshot, GENERATED_AT)
    write_xlsx_report(path, frozen_snapshot, GENERATED_AT, public_export)

    assert_xlsx_has_no_active_or_external_content(path)
    workbook = load_workbook(path, data_only=False)
    assert workbook.sheetnames == SHEET_NAMES

    findings = table_records(
        workbook["Findings"],
        ["Investor", "Expected Value", "Difference", "Computational Status"],
    )
    lp03 = record_matching(findings, "Investor", "LP03")
    lp04 = record_matching(findings, "Investor", "LP04")
    lp06 = record_matching(findings, "Investor", "LP06")
    assert_close_money(record_value(lp03, "Expected Value"), 37_500)
    assert_close_money(record_value(lp03, "Difference"), 12_500)
    assert_close_money(record_value(lp04, "Expected Value"), 40_000)
    assert_close_money(record_value(lp04, "Difference"), 10_000)
    assert record_value(lp06, "Expected Value") is None
    assert record_value(lp06, "Computational Status") == "CANNOT_VERIFY"

    calculations = table_records(
        workbook["Calculations"],
        ["Investor", "Application Formula", "Formula Text", "Server Calculated Expected Fee"],
    )
    lp03_calculation = record_matching(calculations, "Investor", "LP03")
    assert str(record_value(lp03_calculation, "Application Formula")).startswith("=")
    formula_text_cell = workbook["Calculations"]["F10"]
    assert str(record_value(lp03_calculation, "Formula Text")).startswith("=")
    assert formula_text_cell.data_type == "s"
    assert_close_money(record_value(lp03_calculation, "Server Calculated Expected Fee"), 37_500)

    audit = table_records(workbook["Audit Trail"], ["Action", "Run Version", "Note"])
    assert any(record_value(record, "Action") == "OUTPUT_SNAPSHOT_FROZEN" for record in audit)
    assert workbook["Findings"].freeze_panes == "B8"
    assert workbook["Calculations"].auto_filter.ref is not None


def test_untrusted_spreadsheet_content_is_literal_not_formula(
    tmp_path: Path,
    fixture_payload: dict,
) -> None:
    payload = copy.deepcopy(fixture_payload)
    payload["run_id"] = "formula-injection-test"
    payload["findings"][0]["explanation"] = " \t\ufeff=HYPERLINK(\"https://bad.invalid\")"
    payload["evidence_references"][0]["quoted_text"] = "+SUM(1,2)"
    frozen = FileSnapshotStore(tmp_path / "snapshots").freeze(payload)
    path = tmp_path / "injection.xlsx"
    write_xlsx_report(path, frozen, GENERATED_AT, build_review_export(frozen, GENERATED_AT))

    workbook = load_workbook(path, data_only=False)
    formulas = []
    string_cells = []
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.data_type == "f":
                    formulas.append((sheet.title, cell.coordinate, cell.value))
                elif isinstance(cell.value, str):
                    string_cells.append((sheet.title, cell.coordinate, cell.value, cell.data_type))

    assert formulas
    assert all(
        sheet == "Calculations"
        and coordinate.startswith("E")
        and re.fullmatch(r"=B\d+\*C\d+\*D\d+", str(value))
        for sheet, coordinate, value in formulas
    )
    assert any(
        value == "' \t\ufeff=HYPERLINK(\"https://bad.invalid\")" and data_type == "s"
        for _, _, value, data_type in string_cells
    )
    assert any(
        value == "'+SUM(1,2)" and data_type == "s"
        for _, _, value, data_type in string_cells
    )
    assert_xlsx_has_no_active_or_external_content(path)
