from __future__ import annotations

import copy
import json
from datetime import datetime, timezone
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

from openpyxl import load_workbook

from app.relay.contracts import adapt_review_snapshot
from app.relay.json_export import build_review_export, write_json_export
from app.relay.pdf_export import write_pdf_report
from app.relay.snapshot_store import FileSnapshotStore
from app.relay.xlsx_export import write_xlsx_report

from .artifact_assertions import pdf_text, record_matching, record_value, table_records


GENERATED_AT = datetime(2026, 9, 5, 12, 0, tzinfo=timezone.utc)
LARGE_EXACT_AMOUNT = Decimal("90071992547409.01")


def _canonical_with_calculation(fixture_payload: dict, run_id: str) -> tuple[dict, dict, dict]:
    payload = adapt_review_snapshot(copy.deepcopy(fixture_payload)).to_jsonable()
    payload["run_id"] = run_id
    finding = next(item for item in payload["findings"] if item["calculation_id"])
    calculation = next(
        item for item in payload["calculations"]
        if item["calculation_id"] == finding["calculation_id"]
    )
    return payload, finding, calculation


def test_large_cent_amount_stays_exact_in_snapshot_json_pdf_and_xlsx(
    tmp_path: Path,
    fixture_payload: dict,
    export_schema_path: Path,
) -> None:
    payload, finding, calculation = _canonical_with_calculation(
        fixture_payload, "large-exact-money",
    )
    exact = format(LARGE_EXACT_AMOUNT, "f")
    finding.update(
        administrator_value=exact,
        expected_value=exact,
        difference="0.00",
        computational_status="MATCH",
        variance_direction="NO_VARIANCE",
    )
    calculation.update(
        fee_base=exact,
        annual_rate_fraction="1",
        period_factor="1",
        expected_value=exact,
        reported_value=exact,
        difference="0.00",
    )
    frozen = FileSnapshotStore(tmp_path / "snapshots").freeze(payload)

    assert frozen.snapshot.findings[0].administrator_value is not None
    adapted_finding = next(
        item for item in frozen.snapshot.findings if item.finding_id == finding["finding_id"]
    )
    adapted_calculation = next(
        item for item in frozen.snapshot.calculations
        if item.calculation_id == calculation["calculation_id"]
    )
    assert adapted_finding.expected_value == LARGE_EXACT_AMOUNT
    assert adapted_calculation.expected_value == LARGE_EXACT_AMOUNT

    public_export = build_review_export(frozen, GENERATED_AT)
    json_path = tmp_path / "review.json"
    write_json_export(json_path, frozen, GENERATED_AT, export_schema_path)
    parsed = json.loads(json_path.read_text(encoding="utf-8"))
    exported_finding = next(
        item for item in parsed["findings"] if item["finding_id"] == finding["finding_id"]
    )
    exported_calculation = next(
        item for item in parsed["calculations"]
        if item["calculation_id"] == calculation["calculation_id"]
    )
    assert exported_finding["administrator_value"] == exact
    assert exported_finding["expected_value"] == exact
    assert exported_calculation["fee_base"] == exact
    assert exported_calculation["expected_value"] == exact

    pdf_path = tmp_path / "review.pdf"
    write_pdf_report(pdf_path, frozen, GENERATED_AT)
    assert "£90,071,992,547,409.01" in pdf_text(pdf_path).replace("\n", "")

    xlsx_path = tmp_path / "review.xlsx"
    write_xlsx_report(xlsx_path, frozen, GENERATED_AT, public_export)
    workbook = load_workbook(xlsx_path, data_only=False)
    try:
        findings = table_records(workbook["Findings"], ["Investor", "Expected Value"])
        finding_row = record_matching(findings, "Investor", finding["investor_id"])
        calculations = table_records(
            workbook["Calculations"],
            ["Investor", "Fee Base", "Application Formula", "Server Calculated Expected Fee"],
        )
        calculation_row = record_matching(calculations, "Investor", finding["investor_id"])
        assert record_value(finding_row, "Expected Value") == exact
        assert record_value(calculation_row, "Fee Base") == exact
        assert record_value(calculation_row, "Server Calculated Expected Fee") == exact
        assert "omitted to prevent precision loss" in str(
            record_value(calculation_row, "Application Formula")
        )
    finally:
        workbook.close()


def test_xlsx_formula_applies_the_verifiers_round_half_up_cent_rule(
    tmp_path: Path,
    fixture_payload: dict,
) -> None:
    payload, finding, calculation = _canonical_with_calculation(
        fixture_payload, "round-half-up-money",
    )
    finding.update(
        administrator_value="0.02",
        expected_value="0.02",
        difference="0.00",
        computational_status="MATCH",
        variance_direction="NO_VARIANCE",
    )
    calculation.update(
        fee_base="3",
        annual_rate_fraction="0.005",
        period_factor="1",
        expected_value="0.02",
        reported_value="0.02",
        difference="0.00",
    )
    frozen = FileSnapshotStore(tmp_path / "snapshots").freeze(payload)
    path = tmp_path / "rounded.xlsx"
    write_xlsx_report(path, frozen, GENERATED_AT, build_review_export(frozen, GENERATED_AT))

    formulas = load_workbook(path, data_only=False)
    cached = load_workbook(path, data_only=True)
    try:
        formula_rows = table_records(
            formulas["Calculations"],
            ["Investor", "Application Formula", "Server Calculated Expected Fee"],
        )
        cached_rows = table_records(
            cached["Calculations"],
            ["Investor", "Application Formula", "Server Calculated Expected Fee"],
        )
        formula_row = record_matching(formula_rows, "Investor", finding["investor_id"])
        cached_row = record_matching(cached_rows, "Investor", finding["investor_id"])
        independently_recalculated = (
            Decimal("3") * Decimal("0.005") * Decimal("1")
        ).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

        assert str(record_value(formula_row, "Application Formula")).startswith("=ROUND(")
        assert Decimal(str(record_value(cached_row, "Application Formula"))) == Decimal("0.02")
        assert Decimal(str(record_value(cached_row, "Server Calculated Expected Fee"))) == Decimal("0.02")
        assert independently_recalculated == Decimal("0.02")
    finally:
        formulas.close()
        cached.close()
