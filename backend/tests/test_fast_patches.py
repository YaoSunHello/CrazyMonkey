"""Source-bound corrected copies; model calls are unnecessary for these tests."""
from __future__ import annotations

import hashlib
import tempfile
import unittest
from copy import copy
from decimal import Decimal
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from app.atlas.fixtures import generate_synthetic_pack
from app.atlas.ingestion import normalize_file
from app.runtime.challenger import challenge
from app.runtime.contracts import Challenge, NumericInput, Operation, VerificationPlan
from app.runtime.executor import execute
from app.runtime.investigation_evidence import EvidenceStore
from app.runtime.patches import PatchProposal, TRAIL_HEADERS, apply_patches, propose_patch
from app.runtime.planner import offline_plan


class FastPatchTests(unittest.TestCase):
    def setUp(self):
        self.temporary = tempfile.TemporaryDirectory(prefix="verified-patch-test-")
        self.root = Path(self.temporary.name)
        self.sources = self.root / "input"
        generate_synthetic_pack(self.sources)
        self.workbook_path = self.sources / "Administrator_NAV_Q3_2026.xlsx"
        self.output = self.root / "corrected"
        self.refresh()

    def tearDown(self):
        self.temporary.cleanup()

    def refresh(self):
        documents = [normalize_file(path, original_storage_key=str(path))
                     for path in sorted(self.sources.rglob("*"))
                     if path.suffix.lower() in (".xlsx", ".csv", ".pdf")]
        self.store = EvidenceStore(documents)
        self.plan = next(plan for plan in offline_plan(self.store).checks if plan.entity_id == "LP03")
        self.calculation = execute(self.plan, self.store)
        self.review = challenge(self.plan, self.calculation, self.store)

    def proposal(self):
        proposal = propose_patch(self.plan, self.calculation, self.store, self.review)
        self.assertIsNotNone(proposal)
        return proposal

    def test_lp03_corrected_copy_reopens_and_preserves_original_and_other_cells(self):
        workbook = load_workbook(self.workbook_path)
        sheet = workbook["Investor Fees"]
        sheet["H12"] = "=SUM(F4:F9)"
        sheet["H12"].font = Font(bold=True, color="BB2211")
        sheet["F6"].fill = PatternFill("solid", fgColor="FFF2CC")
        workbook.save(self.workbook_path)
        workbook.close()
        self.refresh()
        before = self.workbook_path.read_bytes()
        proposal = self.proposal()
        result = apply_patches([proposal], self.store, self.output)
        corrected_path = Path(result[0]["output_file"])
        self.assertEqual(corrected_path.name, "Administrator_NAV_Q3_2026_FIXED.xlsx")
        self.assertEqual(self.workbook_path.read_bytes(), before)
        self.assertEqual(result[0]["source_hash"], hashlib.sha256(before).hexdigest())
        self.assertEqual(result[0]["source_hash_after"], result[0]["source_hash"])
        self.assertEqual(result[0]["patch_count"], 1)
        original = load_workbook(self.workbook_path)
        corrected = load_workbook(corrected_path)
        try:
            self.assertEqual(corrected["Investor Fees"]["F6"].value, 37500)
            self.assertEqual(corrected["Investor Fees"]["F6"].data_type, "n")
            for source_sheet in original.worksheets:
                target_sheet = corrected[source_sheet.title]
                for source_cell in source_sheet._cells.values():
                    target = target_sheet[source_cell.coordinate]
                    self.assertEqual(copy(source_cell._style), copy(target._style))
                    if (source_sheet.title, source_cell.coordinate) != ("Investor Fees", "F6"):
                        self.assertEqual(target.value, source_cell.value)
                        self.assertEqual(target.data_type, source_cell.data_type)
            trail = corrected["Audit Trail"]
            self.assertEqual(tuple(cell.value for cell in trail[1]), TRAIL_HEADERS)
            self.assertEqual(trail["B2"].value, "Investor Fees")
            self.assertEqual(trail["C2"].value, "F6")
            self.assertEqual(trail["D2"].value, "50000")
            self.assertEqual(trail["E2"].value, "37500.00")
            self.assertEqual(trail["F2"].value, self.plan.check_id)
            self.assertIn(self.plan.inputs[self.plan.reported_input].evidence_id, trail["H2"].value)
            self.assertEqual(corrected["Investor Fees"]["H12"].value, "=SUM(F4:F9)")
        finally:
            original.close()
            corrected.close()

    def test_unaccepted_results_do_not_produce_patches(self):
        for status in ("MATCH", "CANNOT_VERIFY"):
            with self.subTest(status=status):
                self.assertIsNone(propose_patch(self.plan, {**self.calculation, "status": status}, self.store, self.review))
        for status in ("CHALLENGE", "INSUFFICIENT_EVIDENCE"):
            with self.subTest(status=status):
                review = self.review.model_copy(update={"status": status})
                self.assertIsNone(propose_patch(self.plan, self.calculation, self.store, review))
        self.assertIsNone(propose_patch(self.plan, self.calculation, self.store, Challenge(status="PASS", checks={})))

    def test_unknown_evidence_and_forged_calculation_are_rejected(self):
        bad_plan = self.plan.model_copy(deep=True)
        bad_plan.inputs[bad_plan.reported_input].evidence_id = "ev_missing"
        with self.assertRaises(ValueError):
            propose_patch(bad_plan, self.calculation, self.store, self.review)
        with self.assertRaisesRegex(ValueError, "calculation"):
            propose_patch(self.plan, {**self.calculation, "expected": "12345.00"}, self.store, self.review)

    def test_equal_numeric_input_from_another_investor_cannot_authorize_patch(self):
        plan = self.plan.model_copy(deep=True)
        base_name = next(name for name, spec in plan.inputs.items()
                         if name != plan.reported_input and spec.unit == "money")
        other_base = next(ref for ref in self.store.refs.values()
                          if ref.sheet == "Investor Fees" and ref.cell == "C4")
        plan.inputs[base_name].evidence_id = other_base.evidence_id
        calculation = execute(plan, self.store)
        self.assertEqual(calculation["expected"], "37500.00")
        self.assertIsNone(propose_patch(plan, calculation, self.store, self.review))

    def test_verified_csv_discrepancy_has_no_workbook_patch_destination(self):
        csv_path = self.root / "invoice.csv"
        csv_path.write_text("Account ID,Fund,Currency,Quantity,Unit Price,Line Total\nA-42,Demo Fund,GBP,3,10,40\n")
        store = EvidenceStore([normalize_file(csv_path, original_storage_key=str(csv_path))])
        plan = offline_plan(store).checks[0]
        calculation = execute(plan, store)
        review = challenge(plan, calculation, store)
        self.assertEqual(review.status, "PASS")
        self.assertIsNone(propose_patch(plan, calculation, store, review))

    def test_serialized_or_tampered_proposals_have_no_write_authority(self):
        proposal = self.proposal()
        reconstructed = PatchProposal.model_validate(proposal.model_dump())
        changed = proposal.model_copy(update={"new_value": "12345.00"})
        wrong_cell = proposal.model_copy(update={"cell": "F7"})
        for candidate in (reconstructed, changed, wrong_cell):
            with self.subTest(candidate=candidate.cell), self.assertRaisesRegex(ValueError, "verified execution"):
                apply_patches([candidate], self.store, self.output)
        self.assertFalse(self.output.exists())

    def test_evidence_list_mutation_invalidates_private_proof(self):
        proposal = self.proposal()
        proposal.evidence_ids.append("ev_missing")
        with self.assertRaisesRegex(ValueError, "modified"):
            apply_patches([proposal], self.store, self.output)

    def test_changed_original_is_rejected_before_any_output(self):
        proposal = self.proposal()
        workbook = load_workbook(self.workbook_path)
        workbook["Investor Fees"]["F6"] = 50001
        workbook.save(self.workbook_path)
        workbook.close()
        with self.assertRaisesRegex(ValueError, "changed"):
            apply_patches([proposal], self.store, self.output)
        self.assertFalse(self.output.exists())

    def test_existing_output_is_never_replaced(self):
        proposal = self.proposal()
        self.output.mkdir()
        destination = self.output / "Administrator_NAV_Q3_2026_FIXED.xlsx"
        destination.write_bytes(b"existing work")
        with self.assertRaisesRegex(ValueError, "overwrite"):
            apply_patches([proposal], self.store, self.output)
        self.assertEqual(destination.read_bytes(), b"existing work")

    def test_duplicate_and_conflicting_cell_proposals_reject_entire_batch(self):
        first, second = self.proposal(), self.proposal()
        with self.assertRaisesRegex(ValueError, "same cell"):
            apply_patches([first, second], self.store, self.output)
        self.assertFalse(self.output.exists())

    def test_formula_target_cannot_generate_a_proposal(self):
        ref = self.store.get(self.plan.inputs[self.plan.reported_input].evidence_id)
        ref.formula = "=50000"
        ref.cache_status = "MISSING"
        with self.assertRaisesRegex(ValueError, "formula"):
            propose_patch(self.plan, self.calculation, self.store, self.review)

    def test_existing_trail_is_preserved_and_appended(self):
        workbook = load_workbook(self.workbook_path)
        trail = workbook.create_sheet("Audit Trail")
        trail.append(TRAIL_HEADERS)
        trail.append(["previous", "Other", "A1", "10", "20", "old", "Prior change", "old-evidence"])
        workbook.save(self.workbook_path)
        workbook.close()
        self.refresh()
        result = apply_patches([self.proposal()], self.store, self.output)
        corrected = load_workbook(result[0]["output_file"])
        try:
            self.assertEqual(corrected["Audit Trail"]["A2"].value, "previous")
            self.assertEqual(corrected["Audit Trail"]["C3"].value, "F6")
        finally:
            corrected.close()

    def test_ambiguous_existing_trail_rejects_before_writing(self):
        workbook = load_workbook(self.workbook_path)
        workbook.create_sheet("Audit Trail").append(["Unrelated original worksheet"])
        workbook.save(self.workbook_path)
        workbook.close()
        self.refresh()
        with self.assertRaisesRegex(ValueError, "column headers"):
            apply_patches([self.proposal()], self.store, self.output)
        self.assertFalse(self.output.exists())

    def test_untrusted_audit_text_is_not_a_formula(self):
        self.plan.rationale = '=HYPERLINK("https://invalid.example", "untrusted source text")'
        proposal = self.proposal()
        result = apply_patches([proposal], self.store, self.output)
        corrected = load_workbook(result[0]["output_file"])
        try:
            cell = corrected["Audit Trail"]["G2"]
            self.assertEqual(cell.value, self.plan.rationale)
            self.assertEqual(cell.data_type, "s")
        finally:
            corrected.close()

    def test_publication_race_cannot_overwrite_concurrent_output(self):
        proposal = self.proposal()
        def concurrent_output(temporary, destination):
            Path(destination).write_bytes(b"concurrent workbook")
            raise FileExistsError("destination already exists")
        with patch("app.runtime.patches.os.link", side_effect=concurrent_output):
            with self.assertRaises(FileExistsError):
                apply_patches([proposal], self.store, self.output)
        destination = self.output / "Administrator_NAV_Q3_2026_FIXED.xlsx"
        self.assertEqual(destination.read_bytes(), b"concurrent workbook")
        self.assertEqual(list(self.output.iterdir()), [destination])

    def test_excel_precision_and_fractional_pennies_rejected(self):
        data = self.proposal().model_dump()
        for value in ("12345678901234.56", "0.001", "NaN", "Infinity"):
            with self.subTest(value=value), self.assertRaises(ValueError):
                PatchProposal.model_validate({**data, "new_value": value})

    def test_empty_batch_writes_nothing(self):
        self.assertEqual(apply_patches([], self.store, self.output), [])
        self.assertFalse(self.output.exists())

    def test_different_source_directories_with_same_basename_are_rejected(self):
        proposals = []
        documents = []
        for index in (1, 2):
            folder = self.root / f"pack-{index}"
            folder.mkdir()
            path = folder / "Schedule.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Charges"
            sheet.append(["Account ID", "Fund", "Currency", "Quantity", "Unit Price", "Line Total"])
            sheet.append([f"A-{index}", "Demo Fund", "GBP", 3, 10, 40 + index])
            workbook.save(path)
            workbook.close()
            documents.append(normalize_file(path, original_storage_key=str(path)))
        store = EvidenceStore(documents)
        for document in documents:
            refs = {ref.cell: ref for ref in document.evidence}
            plan = VerificationPlan(
                check_id=document.document.document_id, title="Quantity and price", check_type="quantity_price",
                entity_id=refs["A2"].original_value, fund_name="Demo Fund", currency="GBP", rationale="Quantity times price",
                inputs={"quantity": NumericInput(evidence_id=refs["D2"].evidence_id),
                        "price": NumericInput(evidence_id=refs["E2"].evidence_id, unit="money"),
                        "reported": NumericInput(evidence_id=refs["F2"].evidence_id, unit="money")},
                reported_input="reported", operation=Operation(operation="multiply", operands=["quantity", "price"]),
                context_evidence_ids=[ref.evidence_id for ref in refs.values()],
            )
            calculation = execute(plan, store)
            proposal = propose_patch(plan, calculation, store, challenge(plan, calculation, store))
            self.assertIsNotNone(proposal)
            proposals.append(proposal)
        with self.assertRaisesRegex(ValueError, "collide"):
            apply_patches(proposals, store, self.output)
        self.assertFalse(self.output.exists())


if __name__ == "__main__":
    unittest.main()
