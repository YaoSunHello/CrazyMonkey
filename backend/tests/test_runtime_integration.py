"""Real original-file -> ATLAS -> runtime -> human -> RELAY integration checks.

No expected-answer snapshot is supplied to the application. The tests generate
ATLAS's original PDF/XLSX/CSV source pack and independently assert its outcomes.
Email transport is explicitly disabled in every test.
"""

from __future__ import annotations

import hashlib
import io
import json
import tempfile
import unittest
from decimal import Decimal
from email import policy
from email.parser import BytesParser
from pathlib import Path
from unittest.mock import patch

from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from pypdf import PdfReader

from app.atlas import normalize_file
from app.atlas.fixtures import generate_synthetic_pack
from app.relay import api as relay_api
from app.relay.email_delivery import EmailDeliveryService
from app.relay.export_service import ExportService, default_export_service
from app.runtime import api as runtime_api
from app.runtime import beacon
from app.runtime.service import ReviewService


class RuntimeIntegrationTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.sources = tempfile.TemporaryDirectory(prefix="runtime-original-sources-")
        cls.source_directory = Path(cls.sources.name)
        generate_synthetic_pack(cls.source_directory)
        cls.source_paths = sorted(
            path for path in cls.source_directory.iterdir()
            if path.suffix.lower() in {".pdf", ".xlsx", ".csv"}
        )
        cls.documents = [normalize_file(path) for path in cls.source_paths]
        cls.schema_path = default_export_service().schema_path

    @classmethod
    def tearDownClass(cls):
        cls.sources.cleanup()

    def setUp(self):
        self.output = tempfile.TemporaryDirectory(prefix="runtime-integration-output-")
        self.addCleanup(self.output.cleanup)
        self.output_root = Path(self.output.name)
        self.exports = ExportService(self.output_root / "relay", self.schema_path)
        self.reviews = ReviewService(export_service=self.exports)
        self.delivery = EmailDeliveryService(
            transport=None, from_address=None, audit_log=self.output_root / "email-log.jsonl",
        )
        app = FastAPI()
        app.include_router(runtime_api.router)
        app.include_router(beacon.router)
        app.include_router(relay_api.router)
        for target, name, value in (
            (runtime_api, "reviews", self.reviews),
            (relay_api, "service", self.exports),
            (relay_api, "delivery", self.delivery),
        ):
            replacement = patch.object(target, name, value)
            replacement.start()
            self.addCleanup(replacement.stop)
        replacement = patch.object(beacon, "_store", return_value=self.reviews)
        replacement.start()
        self.addCleanup(replacement.stop)
        self.client = TestClient(app)
        self.addCleanup(self.client.close)

    def create(self, case_id="integration-case"):
        return self.reviews.create(
            case_id, "Review management fees, show evidence, and prepare PDF and Excel outputs.",
            self.documents, synthetic=True,
        )

    @staticmethod
    def lp(snapshot, investor_id):
        return next(item for item in snapshot.findings if item.investor_id == investor_id)

    def test_original_sources_produce_verified_snapshot_and_readable_all_format_bundle(self):
        self.assertEqual(len(self.documents), 8)
        for path, document in zip(self.source_paths, self.documents, strict=True):
            self.assertEqual(document.document.document_hash, hashlib.sha256(path.read_bytes()).hexdigest())
        record = self.create()
        self.assertEqual(record.result.mode, "DEMO_FIXTURE")
        self.assertEqual(record.snapshot.mode, "SYNTHETIC_DEMO")
        self.assertEqual(record.snapshot.summary.matches, 3)
        self.assertEqual(record.snapshot.summary.discrepancies, 2)
        self.assertEqual(record.snapshot.summary.cannot_verify, 1)
        lp03 = self.lp(record.snapshot, "LP03")
        self.assertEqual(lp03.reported_value, Decimal("50000"))
        self.assertEqual(lp03.expected_value, Decimal("37500"))
        self.assertEqual(lp03.difference, Decimal("12500"))
        self.assertEqual(lp03.status, "DISCREPANCY")
        self.assertEqual(self.lp(record.snapshot, "LP01").status, "MATCH")
        self.assertEqual(self.lp(record.snapshot, "LP05").status, "MATCH")
        self.assertIsNone(self.lp(record.snapshot, "LP06").expected_value)

        frozen = self.exports.snapshot_store.get(record.result.run_id, 1)
        bundle = self.exports.generate_all(frozen)
        self.assertEqual({item.artifact_type for item in bundle.artifacts}, {"pdf", "xlsx", "json", "eml"})
        for descriptor in bundle.artifacts:
            path = bundle.directory / descriptor.filename
            self.assertEqual(hashlib.sha256(path.read_bytes()).hexdigest(), descriptor.sha256)
            self.assertGreater(path.stat().st_size, 0)

        exported = json.loads(bundle.artifact_path("json").read_text(encoding="utf-8"))
        exported_lp03 = next(item for item in exported["findings"] if item["investor_id"] == "LP03")
        self.assertEqual(exported_lp03["expected_value"], 37500)
        self.assertEqual(exported_lp03["difference"], 12500)
        self.assertEqual(exported_lp03["computational_status"], "DISCREPANCY")
        self.assertEqual(exported["export_metadata"]["snapshot_sha256"], frozen.snapshot_sha256)
        pdf = PdfReader(bundle.artifact_path("pdf"))
        self.assertGreater(len(pdf.pages), 0)
        self.assertIn("LP03", "\n".join(page.extract_text() or "" for page in pdf.pages))
        workbook = load_workbook(bundle.artifact_path("xlsx"), data_only=True)
        try:
            self.assertEqual(workbook.sheetnames, ["Summary", "Findings", "Investor Terms", "Calculations", "Sources", "Audit Trail"])
            lp03_row = next(row for row in workbook["Findings"].iter_rows(values_only=True) if "LP03" in row)
            self.assertIn(37500, lp03_row)
            self.assertIn(12500, lp03_row)
        finally:
            workbook.close()
        draft = BytesParser(policy=policy.default).parsebytes(bundle.artifact_path("eml").read_bytes())
        self.assertIsNone(draft["To"])
        self.assertEqual(len(list(draft.iter_attachments())), 3)
        self.assertEqual(bundle.email_draft["status"], "DRAFT_NOT_SENT")
        self.assertFalse(self.delivery.configured)
        self.assertFalse(self.delivery.audit_log.exists())

    def test_human_review_creates_new_immutable_versions_without_changing_financial_status(self):
        record = self.create()
        run_id = record.result.run_id
        finding_id = self.lp(record.snapshot, "LP03").finding_id
        first = self.exports.snapshot_store.get(run_id, 1)
        original_bytes = first.path.read_bytes()
        reviewed = self.reviews.review(run_id, finding_id, "REVIEWED", "Demo reviewer", "Confirmed discrepancy; request administrator response.")
        self.assertEqual(reviewed.snapshot.version, 2)
        self.assertEqual(self.lp(reviewed.snapshot, "LP03").human_review_state, "REVIEWED")
        self.assertEqual(self.lp(reviewed.snapshot, "LP03").status, "DISCREPANCY")
        self.assertEqual(self.lp(reviewed.snapshot, "LP03").expected_value, Decimal("37500"))
        self.assertEqual(first.path.read_bytes(), original_bytes)
        self.assertEqual(self.exports.snapshot_store.get(run_id, 1).snapshot_sha256, first.snapshot_sha256)
        self.assertNotEqual(self.exports.snapshot_store.get(run_id, 2).snapshot_sha256, first.snapshot_sha256)

        follow_up = self.reviews.review(run_id, finding_id, "NEEDS_FOLLOW_UP", "Demo reviewer", "Administrator response still required.")
        self.assertEqual(follow_up.snapshot.version, 3)
        self.assertEqual(len(follow_up.snapshot.audit_trail), 2)
        self.assertEqual([item.run_version for item in follow_up.snapshot.audit_trail], [2, 3])
        self.assertEqual(follow_up.snapshot.audit_trail[0].note, "Confirmed discrepancy; request administrator response.")
        self.assertEqual(follow_up.snapshot.audit_trail[1].previous_review_state, "REVIEWED")
        self.assertEqual(self.lp(follow_up.snapshot, "LP03").status, "DISCREPANCY")
        previous_snapshot = self.exports.snapshot_store.get(run_id, 2).snapshot
        previous_decision = next(item for item in previous_snapshot.human_review_decisions if item.finding_id == finding_id)
        self.assertEqual(previous_decision.state, "REVIEWED")
        self.assertEqual(len(previous_snapshot.audit_trail), 1)
        self.assertEqual([item.run_version for item in previous_snapshot.audit_trail], [2])
        self.assertEqual([item.run_id for item in previous_snapshot.audit_trail], [run_id])
        latest_snapshot = self.exports.snapshot_store.get(run_id, 3).snapshot
        self.assertEqual([item.run_version for item in latest_snapshot.audit_trail], [2, 3])
        self.assertEqual([item.run_id for item in latest_snapshot.audit_trail], [run_id, run_id])
        beacon_finding = next(
            item for item in beacon.to_beacon(follow_up.snapshot, analyst_mode=follow_up.result.mode)["findings"]
            if item["id"] == finding_id
        )
        self.assertEqual([item["version"] for item in beacon_finding["versions"]], [1, 2, 3])
        self.assertTrue(all(item["expectedValue"]["amount"] == 37500 for item in beacon_finding["versions"]))
        self.assertTrue(
            all(Decimal(str(item["applicableRate"])) == Decimal("1.5") for item in beacon_finding["versions"])
        )

    def test_returned_records_cannot_mutate_captured_evidence_or_stored_results(self):
        returned = self.create()
        run_id = returned.result.run_id
        expected_hash = returned.documents[0].document.document_hash
        expected_explanation = returned.result.findings[0].explanation
        returned.documents[0].document.document_hash = "f" * 64
        returned.result.findings[0].explanation = "Caller-side mutation"
        returned.snapshot.calculations[0].expected_amount = Decimal("1")
        captured = self.reviews.get(run_id)
        self.assertEqual(captured.documents[0].document.document_hash, expected_hash)
        self.assertEqual(captured.result.findings[0].explanation, expected_explanation)
        self.assertNotEqual(captured.snapshot.calculations[0].expected_amount, Decimal("1"))
        captured.snapshot.findings[0].explanation = "Another caller-side mutation"
        self.assertNotEqual(self.reviews.get(run_id).snapshot.findings[0].explanation, "Another caller-side mutation")

    def test_structured_run_and_result_routes_use_real_runtime(self):
        response = self.client.post("/api/cases/structured-case/run", json={
            "user_instruction": "Check management fees and prepare JSON.",
            "normalized_documents": [item.model_dump(mode="json") for item in self.documents],
            "mode": "DEMO_FIXTURE",
        })
        self.assertEqual(response.status_code, 200, response.text)
        result = response.json()
        lp03 = next(item for item in result["findings"] if item["investor_id"] == "LP03")
        self.assertEqual(lp03["status"], "REVIEW_REQUIRED")
        self.assertEqual(Decimal(lp03["expected"]), Decimal("37500"))
        self.assertEqual(result["repair_count"], 1)
        read = self.client.get("/api/cases/structured-case/result")
        self.assertEqual(read.status_code, 200, read.text)
        self.assertEqual(read.json(), result)
        self.assertEqual(self.client.get(f"/api/cases/{result['run_id']}/result").json(), result)

    def test_unconfigured_model_mode_does_not_fall_back(self):
        response = self.client.post("/api/cases/no-model/run", json={
            "user_instruction": "Review fees.", "normalized_documents": [], "mode": "MODEL",
        })
        self.assertEqual(response.status_code, 503)
        self.assertIn("no fallback", response.json()["detail"])
        self.assertEqual(self.client.get("/api/cases/no-model/result").status_code, 404)

    def test_mismatched_normalized_source_hash_is_rejected_before_running(self):
        documents = [item.model_dump(mode="json") for item in self.documents]
        documents[0]["evidence"][0]["document_hash"] = "0" * 64
        response = self.client.post("/api/cases/bad-hash/run", json={
            "user_instruction": "Review fees.", "normalized_documents": documents,
        })
        self.assertEqual(response.status_code, 422, response.text)
        self.assertEqual(self.client.get("/api/cases/bad-hash/result").status_code, 404)

    def test_actual_beacon_upload_routes_produce_source_linked_computed_findings(self):
        uploads = [("files", (path.name, path.read_bytes())) for path in self.source_paths]
        detected = self.client.post(
            "/api/v1/documents/detect",
            data={"client_file_ids": [f"source-{index}" for index in range(len(uploads))]},
            files=uploads,
        )
        self.assertEqual(detected.status_code, 200, detected.text)
        started = self.client.post(
            "/api/v1/reviews", data={"manifest": json.dumps(detected.json())}, files=uploads,
        )
        self.assertEqual(started.status_code, 201, started.text)
        run_id = started.json()["reviewId"]
        result = self.client.get(f"/api/v1/reviews/{run_id}").json()
        self.assertEqual(result["source"], "ATLAS")
        self.assertEqual(result["mode"], "LIVE_OFFLINE")
        self.assertEqual(len(result["documents"]), 8)
        self.assertEqual(len(result["findings"]), 6)
        lp01 = next(item for item in result["findings"] if item["investorId"] == "LP01")
        self.assertEqual(lp01["status"], "MATCH")
        self.assertNotIn("requiredAction", lp01)
        lp03 = next(item for item in result["findings"] if item["investorId"] == "LP03")
        self.assertEqual(lp03["expectedValue"]["amount"], 37500)
        self.assertEqual(lp03["difference"]["amount"], 12500)
        lp06 = next(item for item in result["findings"] if item["investorId"] == "LP06")
        self.assertEqual(lp06["requiredAction"]["documentRole"], "SIDE_LETTER")
        source_evidence_ids = {item.evidence_id for document in self.documents for item in document.evidence}
        self.assertTrue({item["id"] for item in lp03["evidence"]}.issubset(source_evidence_ids))

    def test_beacon_demo_human_review_and_relay_downloads_share_one_snapshot_identity(self):
        started = self.client.post("/api/v1/demo/reviews")
        self.assertEqual(started.status_code, 201, started.text)
        run_id = started.json()["reviewId"]
        self.assertEqual(self.client.get(f"/api/v1/reviews/{run_id}/progress").json()["state"], "COMPLETE")
        review = self.client.get(f"/api/v1/reviews/{run_id}")
        self.assertEqual(review.status_code, 200, review.text)
        payload = review.json()
        self.assertEqual(payload["source"], "ATLAS")
        self.assertEqual(payload["mode"], "SYNTHETIC_DEMO")
        self.assertFalse(payload["outputCapabilities"]["emailSend"])
        lp03 = next(item for item in payload["findings"] if item["investorId"] == "LP03")
        self.assertEqual(lp03["status"], "DISCREPANCY")
        human = self.client.patch(f"/api/v1/reviews/{run_id}/findings/{lp03['id']}/review", json={
            "state": "REVIEWED", "reviewerName": "Demo reviewer", "note": "Discrepancy requires administrator follow-up.",
        })
        self.assertEqual(human.status_code, 200, human.text)
        self.assertEqual(human.json()["status"], "DISCREPANCY")
        self.assertEqual(human.json()["humanReviewState"], "REVIEWED")
        frozen = self.exports.snapshot_store.get(run_id, 2)
        for kind in ("pdf", "excel", "json"):
            response = self.client.get(
                f"/api/v1/reviews/{run_id}/exports/{kind}", params={"version": 2},
            )
            self.assertEqual(response.status_code, 200, response.text if kind == "json" else kind)
            self.assertEqual(response.headers["x-review-version"], "2")
            self.assertEqual(response.headers["x-snapshot-sha256"], frozen.snapshot_sha256)
            if kind == "pdf":
                self.assertGreater(len(PdfReader(io.BytesIO(response.content)).pages), 0)
            elif kind == "json":
                exported_lp03 = next(item for item in response.json()["findings"] if item["investor_id"] == "LP03")
                self.assertEqual(exported_lp03["human_review_status"], "REVIEWED")
                self.assertEqual(exported_lp03["computational_status"], "DISCREPANCY")
        eml = self.client.get(f"/api/runs/{run_id}/versions/2/exports/eml")
        self.assertEqual(eml.status_code, 200)
        draft = BytesParser(policy=policy.default).parsebytes(eml.content)
        self.assertIsNone(draft["To"])
        self.assertEqual(len(list(draft.iter_attachments())), 3)
        prepared = self.client.post(
            f"/api/v1/reviews/{run_id}/email/prepare", params={"version": 2},
        )
        self.assertEqual(prepared.status_code, 200, prepared.text)
        self.assertEqual(prepared.json()["status"], "DRAFT")
        self.assertEqual(prepared.json()["review_version"], 2)
        self.assertEqual(prepared.json()["snapshot_sha256"], frozen.snapshot_sha256)
        self.assertFalse(prepared.json()["send_available"])
        preview = self.client.post(f"/api/runs/{run_id}/email/preview", json={
            "version": 2, "recipient": "reviewer@example.test",
        })
        self.assertEqual(preview.status_code, 200, preview.text)
        self.assertEqual(preview.json()["status"], "PREVIEW_NOT_SENT")
        self.assertEqual(preview.json()["snapshot_sha256"], frozen.snapshot_sha256)
        self.assertFalse(preview.json()["send_configured"])
        self.assertFalse(self.delivery.audit_log.exists())


if __name__ == "__main__":
    unittest.main()
