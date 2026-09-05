"""Independent blind source pair; keep control files away from the investigator.

Run this test-only builder and give the investigator only the printed input_dir.
The sibling control directory is an evaluator oracle, never a runtime input.
All entities, terms and amounts below are fictional.
"""

from __future__ import annotations

import argparse
import hashlib
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Border, Font, PatternFill, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, PageBreak


def _write_workbook(path: Path) -> None:
    """Reproduce the original blind case using the backend's Python libraries."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Account movement"
    sheet.sheet_view.showGridLines = False
    for row in sheet.iter_rows(min_row=1, max_row=18, min_col=1, max_col=5):
        for cell in row:
            cell.font = Font(name="Helvetica", size=11, color="25313F")
    for column, width in {"A": 3, "B": 34, "C": 3, "D": 43, "E": 3}.items():
        sheet.column_dimensions[column].width = width
    sheet["B1"] = "Account movement statement"
    sheet["B1"].font = Font(name="Helvetica", size=16, bold=True, color="25313F")
    sheet["B2"] = "Fictional administrator return. All money in EUR."
    sheet["B2"].font = Font(name="Helvetica", size=11, italic=True, color="596572")
    sheet["B4"] = "Statement field"
    sheet["D4"] = "Recorded value"
    for row in sheet.iter_rows(min_row=4, max_row=4, min_col=2, max_col=4):
        for cell in row:
            cell.fill = PatternFill("solid", fgColor="29415C")
            cell.font = Font(name="Helvetica", size=11, bold=True, color="FFFFFF")
    records = [
        ("Investment vehicle", "Boreal Infrastructure Partners VII"),
        ("Member account", "ST-684"),
        ("Member legal name", "Peregrine Scholarship Trust"),
        ("Assessment window begins", "2027-04-01"),
        ("Assessment window closes", "2027-06-30"),
        ("Settlement denomination", "EUR"),
        ("Assessment pool recorded", 9600000),
        ("Annual tariff recorded", 0.011),
        ("Quarter fraction recorded", 0.25),
        ("Portfolio oversight levy booked", 26400),
        ("Booking basis", "Hard-coded administrator charge"),
        ("Governing papers supplied", "bundle_47.pdf"),
    ]
    for row, (label, value) in enumerate(records, start=5):
        sheet.cell(row, 2, label)
        sheet.cell(row, 4, value)
    for row in range(4, 17):
        sheet.row_dimensions[row].height = 25
    for coordinate, number_format in {
        "D11": "#,##0.00", "D12": "0.00%", "D13": "0.00", "D14": "#,##0.00",
    }.items():
        sheet[coordinate].number_format = number_format
    for cell in sheet[14][1:4]:
        cell.border = Border(top=Side(style="thin", color="B5BDC6"))
        cell.font = Font(name="Helvetica", size=11, bold=True, color="25313F")
    sheet["B18"] = "The portfolio oversight levy is the recurring management fee."
    sheet["B18"].font = Font(name="Helvetica", size=11, italic=True, color="596572")
    workbook.save(path)
    workbook.close()


def _write_document(path: Path) -> None:
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle("Text", fontName="Helvetica", fontSize=11,
                              leading=16, spaceAfter=12))
    styles.add(ParagraphStyle("Section", fontName="Helvetica-Bold", fontSize=12,
                              leading=16, spaceBefore=10, spaceAfter=8,
                              textColor=colors.HexColor("#29415C")))
    story = []

    def heading(value: str) -> None:
        story.append(Paragraph(value, styles["Section"]))

    def paragraph(value: str) -> None:
        story.append(Paragraph(value, styles["Text"]))

    story.append(Paragraph("Boreal Infrastructure Partners VII", styles["Title"]))
    paragraph("Operating agreement and account schedule. Fictional source documents.")
    heading("Article A - Identity and scope")
    paragraph("This agreement governs Boreal Infrastructure Partners VII and its member "
              "Peregrine Scholarship Trust, account reference ST-684. It is effective from "
              "1 January 2027 with no expiry date. The assessment window under review runs "
              "from 1 April 2027 through 30 June 2027, inclusive.")
    heading("Article B - Portfolio oversight levy")
    paragraph("The portfolio oversight levy is the fund's recurring management fee, "
              "charged to each member in euros (EUR). The ordinary annual tariff is "
              "1.60% of the agreed assessment pool. For a full calendar quarter, multiply "
              "the annual tariff by the quarter fraction 0.25 and by the assessment pool.")
    paragraph("The period charge is rounded to the nearest euro cent using round half up. "
              "The reconciliation tolerance is EUR 0.01. No offsets, taxes, rebates, "
              "minimum charges or other adjustments apply to this assessment window.")
    heading("Schedule C - Agreed assessment pool")
    paragraph("For account ST-684, Peregrine Scholarship Trust, the assessment pool is "
              "EUR 9,600,000.00 throughout 1 April 2027 to 30 June 2027. It is the "
              "agreed charge basis for this account and fund; closing account value is "
              "not a substitute for this pool.")
    heading("Article D - Precedence")
    paragraph("An account-specific rider replaces Article B only for the identified "
              "member, the same fund and levy, and only during the rider's stated period "
              "of application. A rider signed before its commencement date has no "
              "effect on earlier assessment windows. The ordinary tariff remains "
              "binding until the rider commences.")
    story.append(PageBreak())
    story.append(Paragraph("Account rider R-28", styles["Title"]))
    paragraph("Supplement to the operating agreement. Fictional source document.")
    heading("R-28.1 - Parties and relationship")
    paragraph("Boreal Infrastructure Partners VII grants the following portfolio "
              "oversight levy concession solely to Peregrine Scholarship Trust, "
              "account ST-684. This rider supplements the operating agreement on "
              "page 1 and concerns the recurring management fee defined in Article B.")
    heading("R-28.2 - Signature and commencement")
    paragraph("Signed on 15 March 2027. The reduced annual tariff of 1.10% commences "
              "on 1 July 2027, with no stated end date. Signature does not accelerate "
              "commencement. This concession does not apply to any charge for a period "
              "ending before 1 July 2027.")
    heading("R-28.3 - Unchanged provisions")
    paragraph("The assessment-pool definition, full-quarter fraction, rounding, "
              "currency and reconciliation tolerance remain as stated in the operating "
              "agreement. No other member-specific terms exist for this account.")

    def footer(page, document) -> None:
        page.setFont("Helvetica", 8)
        page.setFillColor(colors.HexColor("#596572"))
        page.drawString(20 * mm, 15 * mm, "Fictional source pack - no real investor or fund")
        page.drawRightString(A4[0] - 20 * mm, 15 * mm, f"Page {document.page}")

    SimpleDocTemplate(str(path), pagesize=A4, rightMargin=20 * mm,
                      leftMargin=20 * mm, topMargin=18 * mm, bottomMargin=24 * mm,
                      title="Boreal account governing papers", author="Independent QA",
                      invariant=1).build(story, onFirstPage=footer, onLaterPages=footer)


def generate_blind_pair(output_dir: Path) -> dict:
    output_dir = output_dir.resolve()
    inputs = output_dir / "input"
    control = output_dir / "control"
    inputs.mkdir(parents=True, exist_ok=False)
    control.mkdir(parents=True, exist_ok=False)
    _write_workbook(inputs / "parcel_82.xlsx")
    _write_document(inputs / "bundle_47.pdf")
    expected = {
        "synthetic": True,
        "investor_id": "ST-684",
        "investor_name": "Peregrine Scholarship Trust",
        "fund_name": "Boreal Infrastructure Partners VII",
        "status": "DISCREPANCY",
        "currency": "EUR",
        "fee_base": "9600000.00",
        "annual_rate": "0.016",
        "period_factor": "0.25",
        "expected_amount": "38400.00",
        "reported_amount": "26400.00",
        "difference_reported_minus_expected": "-12000.00",
        "reason": "The administrator applied the 1.10% rider before its 1 July 2027 commencement; the quarter ends on 30 June 2027.",
        "source_anchors": {
            "reported": "parcel_82.xlsx / Account movement!D14",
            "recorded_rate": "parcel_82.xlsx / Account movement!D12",
            "base": "parcel_82.xlsx / Account movement!D11; bundle_47.pdf page 1 Schedule C",
            "default_rate_factor_rounding": "bundle_47.pdf page 1 Article B",
            "scope_period_identity": "bundle_47.pdf page 1 Article A",
            "precedence": "bundle_47.pdf page 1 Article D",
            "rider_dates": "bundle_47.pdf page 2 R-28.2",
        },
    }
    (control / "oracle.json").write_text(json.dumps(expected, indent=2) + "\n")
    hashes = {path.name: hashlib.sha256(path.read_bytes()).hexdigest()
              for path in sorted(inputs.iterdir())}
    (control / "input_hashes.json").write_text(json.dumps(hashes, indent=2) + "\n")
    return {"input_dir": str(inputs), "files": sorted(hashes),
            "sha256": hashes, "synthetic": True}


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()
    print(json.dumps(generate_blind_pair(args.output), indent=2))


if __name__ == "__main__":
    main()
