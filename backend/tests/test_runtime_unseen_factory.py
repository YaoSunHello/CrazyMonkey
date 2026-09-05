"""Guard the blind fixture's input/oracle separation and immutable source pack."""

from __future__ import annotations

import hashlib
import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook
from pypdf import PdfReader

from runtime_unseen_factory import generate_blind_pair


class RuntimeUnseenFactoryTests(unittest.TestCase):
    def test_inputs_exclude_oracle_and_cannot_be_overwritten(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary) / "pack"
            manifest = generate_blind_pair(root)
            inputs = Path(manifest["input_dir"])
            self.assertEqual({p.name for p in inputs.iterdir()},
                             {"bundle_47.pdf", "parcel_82.xlsx"})
            self.assertEqual({p.name for p in (root / "control").iterdir()},
                             {"oracle.json", "input_hashes.json"})
            oracle = json.loads((root / "control/oracle.json").read_text())
            workbook = load_workbook(inputs / "parcel_82.xlsx")
            try:
                self.assertEqual(workbook.sheetnames, ["Account movement"])
                self.assertEqual(workbook.active["D6"].value, oracle["investor_id"])
                self.assertEqual(workbook.active["D16"].value, "bundle_47.pdf")
            finally:
                workbook.close()
            self.assertEqual(len(PdfReader(inputs / "bundle_47.pdf").pages), 2)
            with self.assertRaises(FileExistsError):
                generate_blind_pair(root)
            self.assertEqual(
                {p.name: hashlib.sha256(p.read_bytes()).hexdigest() for p in inputs.iterdir()},
                manifest["sha256"],
            )


if __name__ == "__main__":
    unittest.main()
